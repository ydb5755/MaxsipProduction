from datetime import datetime, timedelta
import time
from sqlalchemy import create_engine, MetaData, Table, select
from openpyxl import load_workbook
from xlsxwriter import Workbook
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


def new_telgoo_report():
    fields = ['ENROLLMENT_ID', 'AGENT_NAME', 'PLAN', 'ACCOUNT_STATUS', 'DISCONNECT_REASON', 'MDN']
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'AGENT_NAME', 'PLAN', 'ACCOUNT_STATUS', 'DISCONNECT_REASON', 'MDN'])
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'NEW_CONTACTS')).all()
    print(f'Number of new contacts to be written to report: {len(contacts)}')
    logs.append(f'Number of New Rows: {len(contacts)}')

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-New_Telgoo_Records.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "New_Telgoo_Records")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("New_Telgoo_Records")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "New_Telgoo_Records" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished New Record Report')
    return remote, 'New_Telgoo_Records'

def acp_status_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'DISCONNECT_REASON')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'DISCONNECT_REASON')).all()
    
    contact_len = len(contacts)
    print(f'Number of AcpStatus changes to be written to report: {contact_len}')
    report_summary['ACP_Status'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Telgoo_ACP_Status_Changes.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "AcpStatus")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            # if cell_num == fields.index('Activation_Date'):
            #     cell = cell[:-1]
            #     cell = datetime.strptime(cell, '%Y-%m-%d %H:%M:%S.%f').strftime('%Y-%m-%d %H:%M:%S')
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Telgoo_ACP_Status_Changes")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Telgoo_ACP_Status_Changes" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished AcpStatus Report')
    return remote, 'Acp_Status'

def mdn_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'MDN')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'MDN')).all()
    contact_len = len(contacts)
    print(f'Number of MDN changes to be written to report: {contact_len}')
    report_summary['MDN'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Telgoo_MDN_Changes.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "MDN")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Telgoo_MDN_Changes")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Telgoo_MDN_Changes" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished MDN Report')
    return remote, 'MDN'

def esn_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'ESN')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'ESN')).all()
    contact_len = len(contacts)
    print(f'Number of ESN changes to be written to report: {contact_len}')
    report_summary['Sim_Swap'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Telgoo_Sim_Swaps.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "ESN")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Telgoo_Sim_Swaps")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Telgoo_Sim_Swaps" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished ESN Report')
    return remote, 'Sim_Swaps'

def device_id_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'DEVICE_ID')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'DEVICE_ID')).all()
    contact_len = len(contacts)
    print(f'Number of DEVICE_ID changes to be written to report: {contact_len}')
    report_summary['DEVICE_ID'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Device_Id_Changes.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "Device_Id")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Device_Id_Changes")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Device_Id_Changes" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished Device_Id Report')
    return remote, 'Device_Id'

def make_model_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'MAKE_MODEL')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'MAKE_MODEL')).all()
    contact_len = len(contacts)
    print(f'Number of Make_Model changes to be written to report: {contact_len}')
    report_summary['Make_Model'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Make_Model_Changes.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "Make Model")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Make_Model_Changes")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Make_Model_Changes" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished Make Model Report')
    return remote, 'Make Model'

def product_type_report():
    fields = ['ENROLLMENT_ID', 'Old_Value', 'New_Value']
    # with engine.connect() as conn:
    #     contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
    #                             .where(telgooChanges.c['Change_Date'] == yesterday_slashes_mdY, telgooChanges.c['Field_Name'] == 'PRODUCT_TYPE')).all()
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'Old_Value', 'New_Value'])\
                                .where(telgooChanges.c['Change_Date'] == yesterday, telgooChanges.c['Field_Name'] == 'PRODUCT_TYPE')).all()
    contact_len = len(contacts)
    print(f'Number of Product Type changes to be written to report: {contact_len}')
    report_summary['Product_Type'] = contact_len

    # Build Excel Workbook report
    os.chdir(path_to_reports)
    local_filename = today + '-Product_Type_Changes.xlsx'
    wb = Workbook(local_filename)
    ws = wb.add_worksheet()
    ws.write(0,0, "Device Type")
    ws.write(0,1, 'Total Number of Rows ->')
    ws.write(0,2, len(contacts))
    for field_num, field in enumerate(fields):
        ws.write(1, field_num, field)
    for contact_num, contact in enumerate(contacts):
        for cell_num, cell in enumerate(contact):
            ws.write(contact_num + 2, cell_num, cell)
    wb.close()
    os.chdir("Product_Type_Changes")
    if year not in os.listdir():
        os.mkdir(year, mode=0o777)
        os.chmod(year, mode=0o777)
        os.chdir(year)
    else:
        os.chdir(year)
    if month not in os.listdir():
        os.mkdir(month, mode=0o777)
        os.chmod(month, mode=0o777)
        os.chdir(month)
    else:
        os.chdir(month)
    
    remote = path_to_reports + '/' + "Product_Type_Changes" + '/' + year + '/' + month + '/' + local_filename
    os.chdir(path_to_reports)
    os.rename(local_filename, remote)
    print('Finished Product Type Report')
    return remote, 'Product Type'

def send_email_report(emails_to_send_to: list[str], subject: str = '', body: str = None, attachments:list[MIMEBase] = None):
    for email in emails_to_send_to:
        # Set up the email details
        sender = 'ybaum@maxsiptel.com'  
        receiver = email

        # Create the email message
        message = MIMEMultipart()
        message['From'] = sender
        message['To'] = receiver
        message['Subject'] = subject

        # Add the report content to the email body
        message.attach(MIMEText(body, 'plain'))

        # Add attachments if exist
        if attachments:
            for attachment in attachments:
                message.attach(attachment)

        # Connect to SES SMTP server and send the email
        try:
            smtp_obj = smtplib.SMTP('email-smtp.us-east-2.amazonaws.com', 587)
            smtp_obj.starttls() 
            smtp_obj.login("AKIAWCL4EM4HHB2O5THX", "BNcJaQoFMaQoinh2o6+IGlNl76CMvpz+xbpDJJhpU48q")  
            smtp_obj.sendmail(sender, receiver, message.as_string())
            smtp_obj.quit()
            print("Email sent successfully.")
        except smtplib.SMTPException as e:
            print(f"Error sending email: {str(e)}")

def compile_reports(wb_paths_and_names):
    file_name = 'Daily_Telgoo_Data_Report_' + today + '.xlsx'
    inclusive_workbook = Workbook(file_name)
    for wb, name in wb_paths_and_names:
        ws = inclusive_workbook.add_worksheet(name)
        current = load_workbook(wb, read_only=True)
        current_ws = current.active
        for row_num, row in enumerate(current_ws.rows):
            for cell_num, cell in enumerate(row):
                ws.write(row_num, cell_num, cell.value)
    inclusive_workbook.close()
    
    # Attach excel report
    part = MIMEBase('application', "octet-stream")
    name = inclusive_workbook.filename
    part.set_payload(open(name, "rb").read())
    encoders.encode_base64(part)
    part.add_header('content-disposition', f'attachment; filename={name}')
    return part, name

if __name__ == '__main__':
    path_to_db = 'sqlite:////home/ubuntu/MaxsipReports/instance/site.db'
    path_to_reports = '/home/ubuntu/MaxsipReports/app/static/TelgooReports'
    path_to_csv = '/home/ubuntu/MaxsipReports/pythonFiles/previous_day_file.csv'


    # Establish a database connection
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    telgooChanges = Table("Telgoo_Changes", metadata_obj, autoload_with=engine)

    yesterday = time.mktime((datetime.today() - timedelta(days=2)).date().timetuple())

    # today = datetime.today().strftime('%m-%d-%Y')
    today = '03-12-2024'


    year = datetime.today().strftime('%Y')
    month = datetime.today().strftime('%B')

    
    logs = []
    report_summary = {}

    function_list = [acp_status_report, new_telgoo_report, mdn_report, esn_report, device_id_report, make_model_report, product_type_report]
    wb_paths_and_names = [f() for f in function_list]

    part, name = compile_reports(wb_paths_and_names)

    send_email_report(emails_to_send_to=['yisroel.d.baum@gmail.com', 'cbaum@maxsiptel.com', 'cheskib@gmail.com', 'yschwerd@maxsiptel.com'], #'yschwerd@maxsiptel.com', 'nrotundo@maxsiptel.com', 'nickjrotundo@gmail.com'], 
                      subject='Telgoo Database Update Report',
                      body=f"Telgoo Logs: \n"+ '\n'.join(logs),# + '\nReport Summary:\n',
                      attachments=[part])
    os.remove(name)
    

