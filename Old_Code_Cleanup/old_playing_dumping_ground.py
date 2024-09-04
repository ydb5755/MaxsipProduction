from datetime import datetime, timedelta
from sqlalchemy import create_engine, MetaData, Table, select, update, bindparam, text, insert, distinct, case
import csv
from xlsxwriter import Workbook
import time
import pandas as pd
from openpyxl import load_workbook
from dateutil.parser import parse
import os
import shutil
import re
import requests
import paramiko
from sqlalchemy.sql.expression import func
import json
import zipfile
import psutil




def updating_dates():
    ## Need a list of dicts with an id and a field to update
    table_and_columns_dict = {
    # 'Telgoo_Changes' : ['Change_Date'],
    # 'Unavo_Changes': ['Change_Date', 'Activation_Date']
    # 'customerInfo': ['ActivationDate', 'DeActivationDate', 'CreatedDate', 'FollowupPeriod_CreatedDate']
    'customerInfoTelgoo': ['DEACTIVATION_DATE']
    # 'Device_Qty': ['Date_Of_First_Device', 'Date_Of_Last_Device']
    }
    list_of_row_dicts = []
    with engine.connect() as conn:
        # Create New Column
        # conn.execute(text('ALTER TABLE Telgoo_Changes ADD COLUMN New_Change_Date Integer'))
        for k, v in table_and_columns_dict.items():
            print(f'Starting Table {k}')
            table_name = k
            table_obj = Table(k, metadata_obj, autoload_with=engine)
            for col in v:
                list_of_row_dicts = []
                new_col = f'New_{col}'
                print(f'Starting col: {col}')

                #Update New column with dates converted to timestamp from old col
                contacts = conn.execute(select(table_obj.c['ENROLLMENT_ID', col])).all()
                print('Selection Completed')
                for contact_num, contact in enumerate(contacts):
                    if contact_num % 10000 == 0:
                        print(contact_num)
                    row_dict = {
                        'b_id': contact[0],
                        new_col : ''
                    }
                    if contact[1]:
                        try:
                            row_dict[new_col] = parse(contact[1]).timestamp()
                        except:
                            print(contact)
                            row_dict[new_col] = ''
                    list_of_row_dicts.append(row_dict)

                update_stmt = update(table_obj).where(table_obj.c['ENROLLMENT_ID'] == bindparam('b_id')).values(**{new_col : bindparam(new_col)})
                conn.execute(update_stmt, list_of_row_dicts)
                conn.commit()
                print('Update Completed')

                # Drop old Col
                drop_column_text = f'ALTER TABLE {table_name} DROP COLUMN {col}'
                conn.execute(text(drop_column_text))
                # conn.execute(text('ALTER TABLE Unavo_Changes DROP COLUMN Activation_Date'))
                conn.commit()
                print('Column Dropped')

                # Rename New col to match old col name
                rename_column_text = f'ALTER TABLE {table_name} RENAME COLUMN {new_col} TO {col}'
                conn.execute(text(rename_column_text))
                # conn.execute(text('ALTER TABLE Unavo_Changes RENAME COLUMN New_Activation_Date TO Activation_Date'))
                conn.commit()
                print('Column Renamed')

def checking_dates():
    yesterday = time.mktime((datetime.today() - timedelta(days=4)).date().timetuple())
    print(yesterday)
    


def enrollsWithNoSubs():
    with engine.connect() as conn, open('enrollNoSub.csv', 'w') as csvwritefile:
        csvwriter = csv.writer(csvwritefile)
        contacts = conn.execute(select(main_customer_view).where(main_customer_view.c['NLAD_Subscriber_ID'] == '')).all()
        print(len(contacts))
        key_indexes = {k:v for v, k in enumerate(main_customer_view.c.keys())}
        csvwriter.writerow(main_customer_view.c.keys())
        for contact in contacts:
            data_to_write = [x for x in contact]
            for k, v in key_indexes.items():
                if k in ['DOB', 'Created_Date', 'Order_Date', 'Activation_Date']:
                    if data_to_write[v]:
                        data_to_write[v] = datetime.fromtimestamp(float(data_to_write[v])).strftime('%Y-%m-%d')
            csvwriter.writerow(data_to_write)
            

def getting_subs():
    with engine.connect() as conn:
        # enrollment_id_and_total_payments_dict = {}
        # contacts = conn.execute(select(AgentSalesReports.c['Enrollment_ID', 'Amount'])).all()
        imei_changes_telgoo = conn.execute(select(telgooChanges.c['ENROLLMENT_ID']).where(telgooChanges.c['Field_Name'] == 'DEVICE_ID')).all()
        imei_changes_unavo = conn.execute(select(unavoChanges.c['OrderId']).where(unavoChanges.c['Field_Name'] == 'IMEI')).all()
        total_imei_changes = [x[0] for x in imei_changes_telgoo] + [x[0] for x in imei_changes_unavo]
        print(total_imei_changes[0])
        
def practice():
    wb = load_workbook('Transfer Out Agent Analysis from Telgoo Only 2-28-24.xlsx')
    ws = wb.active
    list_of_enroll_ids = []
    for row_num, row in enumerate(ws.rows):
        if row_num == 0:
            continue
        row_data = row[5].value
        list_of_enroll_ids.append(row_data)
    print(len(list_of_enroll_ids))
    with engine.connect() as conn:
        contacts = conn.execute(select(telgooChanges.c['Change_Date'])
                                .where(telgooChanges.c['Field_Name'] == 'NEW_CONTACTS', 
                                       telgooChanges.c['ENROLLMENT_ID'].in_(list_of_enroll_ids))
                                       ).all()
        print(len(contacts))

def make_copy_of_db():
    shutil.copyfile('/home/ubuntu/MaxsipReports/instance/site.db', f'/home/ubuntu/MaxsipReports/instance/temp_backup.db')
    print('copy made')


def sorting_files():
    files = os.listdir('/home/ubuntu/MaxsipReports/pythonFiles/telgooReports/Adding_New')
    
    # Define a regular expression pattern to extract the numeric part
    pattern = re.compile(r'AO-(\d+)')

    # Sort the files based on the numeric part before the date range
    sorted_files = sorted(files, key=lambda file_name: int(pattern.search(file_name).group(1)))

    # Display the sorted files
    for file in sorted_files:
        print(file)


def parsing_toys():
    #
    # datestring = '5/22/2022  5:41:53'
    # dt = datetime.strptime(parse(datestring).strftime('%m-%d-%Y'), '%m-%d-%Y').timestamp()
    # print(dt)
    with engine.connect() as conn:
        stmt = select(main_customer_view.c['Database', 'City', 'State'])
        print(stmt.selected_columns.keys())


def ntor_generation(start_date: int, end_date: int, user_id:int):
    def example_data():
        a = {
           'id':{
               'data_to_write': {
                   'col':'val',
                   'col':'val',
                   'col':'val',
                   'col':'val'
               },
               'data_from_db':{
                   'col':'val',
                   'col':'val',
                   'col':'val',
                   'col':'val'
               }
           } 
        }
    engine = create_engine('sqlite:////home/ubuntu/MaxsipReports/instance/site.db')
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Transfer_Outs = Table("Transfer_Outs", metadata_obj, autoload_with=engine)
    main_customer_view = Table("Main_Customer_Database", metadata_obj, autoload_with=engine)
    telgooChanges = Table("Telgoo_Changes", metadata_obj, autoload_with=engine)
    # Columns needed from Transfer Out- TRANSACTION_DATE, ACTIVATION_DATE, MASTER_AGENT, EMPLOYEE, ENROLL_ID
    # Columns needed from main_customer_view- Created_Date, Order_Date, Activation_Date, Device_Type, IMEI, NLAD_Subscriber_ID

    # start_date = time.mktime(start_date.timetuple())
    # end_date = time.mktime(end_date.timetuple())
    # user = User.query.filter_by(id=user_id).first()

    contacts_dict = {}
    enrollment_ids = set()
    with engine.connect() as conn:
        stmt = select(Transfer_Outs.c['id', 'TRANSACTION_DATE', 'ACTIVATION_DATE', 'MASTER_AGENT', 'EMPLOYEE', 'ENROLL_ID']).where(Transfer_Outs.c['TRANSACTION_DATE'] >= start_date, Transfer_Outs.c['TRANSACTION_DATE'] <= end_date)
        db_cols = {k:v for v, k in enumerate(stmt.selected_columns.keys())}
        transfers = conn.execute(stmt).all()
    day_ranges = {'0-30':[], '31-60':[], '61-90':[], '91-120':[], '121-360':[], '361-600':[]}
    for contact in transfers:
        if not contact[db_cols['ACTIVATION_DATE']] or not contact[db_cols['TRANSACTION_DATE']]:
            continue
        enrollment_ids.add(contact[db_cols['ENROLL_ID']])
        contacts_dict[contact[db_cols['id']]] = {
                'id':contact[db_cols['id']],
                'Employee Master:Agent':f"{contact[db_cols['MASTER_AGENT']]}:{contact[db_cols['EMPLOYEE']]}",
                'Days in Effect':(float(contact[db_cols['TRANSACTION_DATE']]) - float(contact[db_cols['ACTIVATION_DATE']])) / 86400,
                'Day Range':'>600',
                'ENROLL ID': contact[db_cols['ENROLL_ID']],
                'Created Date':'',
                'Order Date':'',
                'Activation Date DB':'',
                'Device Type':'',
                'IMEI Changes':'',
                'Subscriber ID':'',
                'TRANSACTION DATE': contact[db_cols['TRANSACTION_DATE']],
                'ACTIVATION DATE': contact[db_cols['ACTIVATION_DATE']],
                'MASTER AGENT': contact[db_cols['MASTER_AGENT']],
                'EMPLOYEE': contact[db_cols['EMPLOYEE']]
        }
        days_in_effect = int(contacts_dict[contact[db_cols['id']]]['Days in Effect'])
        if days_in_effect <= 30:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '0-30'
        elif 30 < days_in_effect < 61:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '31-60'
        elif 60 < days_in_effect < 91:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '61-90'
        elif 90 < days_in_effect < 121:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '91-120'
        elif 120 < days_in_effect < 361:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '121-360'
        elif 360 < days_in_effect < 601:
            contacts_dict[contact[db_cols['id']]]['Day Range'] = '361-600'
    print(len(contacts_dict.keys()))
    print(len(enrollment_ids))
    with engine.connect() as conn:
        contacts = conn.execute(select(main_customer_view.c['Customer_Order_Id', 'Created_Date', 'Order_Date', 'Activation_Date', 'Device_Type', 'NLAD_Subscriber_ID'])
                                .where(main_customer_view.c['Customer_Order_Id'].in_(enrollment_ids))).all()
        imei_changes_by_enroll_id = conn.execute(select(telgooChanges.c['ENROLLMENT_ID'])
                                                .where(telgooChanges.c['Field_Name'] == 'DEVICE_ID', telgooChanges.c['ENROLLMENT_ID'].in_(enrollment_ids))).all()
        acp_changes_by_enroll_id = conn.execute(select(telgooChanges.c['ENROLLMENT_ID', 'New_Value', 'Change_Date'])
                                                .where(telgooChanges.c['Field_Name'] == 'DISCONNECT_REASON', 
                                                       telgooChanges.c['ENROLLMENT_ID'].in_(enrollment_ids),
                                                       telgooChanges.c['Change_Date'] < end_date)).all()

    # IMEI
    imei_change_count_dict = {}
    for imei_contact in imei_changes_by_enroll_id:
        if imei_change_count_dict.get(imei_contact[0], None):
            imei_change_count_dict[imei_contact[0]] += 1
        else:
            imei_change_count_dict[imei_contact[0]] = 1
    
    # ACP
    acp_change_dict = {}
    for acp_contact in acp_changes_by_enroll_id:
        row_data = [acp_contact[1], acp_contact[2]]
        if row_data[1]:
            row_data[1] = datetime.fromtimestamp(float(row_data[1])).strftime('%Y-%m-%d')
        if row_data[0] == '':
            row_data[0] = 'Enrolled'
        if acp_change_dict.get(acp_contact[0], None):
            acp_change_dict[acp_contact[0]] += [row_data[0], row_data[1]]
        else:
            acp_change_dict[acp_contact[0]] = [row_data[0], row_data[1]]

    # Details from contacts and join together with imei and acp
    data_for_insert_from_db = {}
    for contact in contacts:
        if imei_change_count_dict.get(contact[0], None):
            imei_changes = imei_change_count_dict[contact[0]]
        else:
            imei_changes = 0
        if acp_change_dict.get(contact[0], None):
            acp_data = acp_change_dict[contact[0]]
        else:
            acp_data = []
        
        data_for_insert_from_db[contact[0]] = {
            'Created Date':contact[1],
            'Order Date':contact[2], 
            'Activation Date':contact[3], 
            'Device Type':contact[4],
            'IMEI Changes': imei_changes,
            'Subscriber ID':contact[5],
            'ACP_Data':acp_data
        }
    for id, contact_dict in contacts_dict.items():
        if data_for_insert_from_db.get(contact_dict['ENROLL ID'], None):
            db_data_contact = data_for_insert_from_db[contact_dict['ENROLL ID']]
            contact_dict['Created Date'] = db_data_contact['Created Date']
            contact_dict['Order Date'] = db_data_contact['Order Date']
            contact_dict['Activation Date DB'] = db_data_contact['Activation Date']
            contact_dict['Device Type'] = db_data_contact['Device Type']
            contact_dict['IMEI Changes'] = db_data_contact['IMEI Changes']
            contact_dict['Subscriber ID'] = db_data_contact['Subscriber ID']


    with Workbook('experiment.xlsx') as wb:
        ws = wb.add_worksheet()
        cols = contacts_dict[list(contacts_dict.keys())[0]].keys()
        ws.write_row(0,0,cols)
        counter = 1
        for id, id_dict in contacts_dict.items():
            row_data = []
            for k, v in id_dict.items():
                if k in ['TRANSACTION DATE', 'ACTIVATION DATE', 'Created Date', 'Order Date', 'Activation Date DB']:
                    if v:
                        v = datetime.fromtimestamp(float(v)).strftime('%Y-%m-%d')
                row_data.append(v)
            ws.write_row(counter,0,row_data)
            counter += 1


def listing_dirs():
    seven_days_ago = time.mktime((datetime.today() - timedelta(days=7)).date().timetuple())
    enroll_ids = ['WMAX14108941', 'WMAX14132869', 'WMAX14147435', 'WMAX14087529', 'WMAX14147593', 'WMAX14149707', 'WMAX14072235', 'WMAX14092459', 'WMAX14094591']
    with engine.connect() as conn:
        for id in enroll_ids:
            stmt = select(telgooChanges.c['ENROLLMENT_ID', 'Change_Date', 'Field_Name', 'Old_Value', 'New_Value']).where(telgooChanges.c['ENROLLMENT_ID'] == id, telgooChanges.c['Change_Date'] > seven_days_ago)
            stmt_columns = stmt.selected_columns.keys()
            imei_changes = conn.execute(stmt).all()
            for record in imei_changes:
                record_with_cols = {k:v for k, v in zip(stmt_columns, record)}
                print(record_with_cols)
            print('\n')

def check_lists():
    list_one = ['CUSTOMER_ID', 'ENROLLMENT_ID', 'FIRST_NAME', 'LAST_NAME', 'ADDRESS1', 'ADDRESS2', 'CITY', 'STATE', 'ZIP', 'SSN', 'IS_HOUSEHOLD', 'IS_TRIBAL', 'IS_SHELTER', 'IS_STATE_DB', 'ESN', 'MDN', 'EMAIL', 'ALTERNATE_PHONE_NUMBER', 'DEVICE_ID', 'MAKE_MODEL', 'NLAD_SUBSCRIBER_ID', 'STATE_DB_MATCHED', 'RECORD_KEEPING', 'ORDER_STATUS', 'PLAN', 'PRODUCT_TYPE', 'TABLET_SUBSIDY_QUALIFICATION', 'CARRIER', 'QUALIFYING_PROGRAM', 'SOURCE', 'AUTHORIZE_CODE', 'MASTER_AGENT_NAME', 'DISTRIBUTOR_AGENT_NAME', 'RETAILER_AGENT_NAME', 'AGENT_NAME', 'DL_NUMBER', 'WINBACK', 'ACCOUNT_STATUS', 'PAYMENT_TYPE', 'SPONSOR_ID', 'CURRENT_APS', 'BQP_FIRST_NAME', 'BQP_LAST_NAME', 'PORTIN_STATUS', 'AGENT_LOGIN_ID', 'DISCONNECT_REASON', 'NLAD_ENROLLMENT_TYPE', 'ENROLLMENT_TYPE', 'CONSENT_FORM_AVAILABE', 'ID_PROOF_AVAILABE', 'DEVICE_EXPECTED_RATE', 'REVIEWED_BY', 'SAME_MONTH_DISCONNECTION', 'ACTIVE_PERIOD', 'CREATED_DATE', 'ORDER_DATE', 'ACTIVATION_DATE', 'DOB', 'BQP_BIRTH_DATE', 'DEVICE_REIMBURSEMENT__DATE', 'DEACTIVATION_DATE']
    list_two = ['CREATED_DATE', 'ORDER_DATE', 'ACTIVATION_DATE', 'CUSTOMER_ID', 'ENROLLMENT_ID', 'FIRST_NAME', 'LAST_NAME', 'ADDRESS1', 'ADDRESS2', 'CITY', 'STATE', 'ZIP', 'SSN', 'DOB', 'IS_HOUSEHOLD', 'IS_TRIBAL', 'IS_SHELTER', 'IS_STATE_DB', 'ESN', 'MDN', 'EMAIL', 'ALTERNATE_PHONE_NUMBER', 'DEVICE_ID', 'MAKE_MODEL', 'NLAD_SUBSCRIBER_ID', 'STATE_DB_MATCHED', 'RECORD_KEEPING', 'ORDER_STATUS', 'PLAN', 'PRODUCT_TYPE', 'CARRIER', 'TABLET_SUBSIDY_QUALIFICATION', 'QUALIFYING_PROGRAM', 'SOURCE', 'AUTHORIZE_CODE', 'MASTER_AGENT_NAME', 'DISTRIBUTOR_AGENT_NAME', 'RETAILER_AGENT_NAME', 'AGENT_NAME', 'DL_NUMBER', 'WINBACK', 'ACCOUNT_STATUS', 'PAYMENT_TYPE', 'SPONSOR_ID', 'CURRENT_APS', 'PORTIN_STATUS', 'AGENT_LOGIN_ID', 'DISCONNECT_REASON', 'DEACTIVATION_DATE', 'MEDICAID_ID', 'BQP_FIRST_NAME', 'BQP_LAST_NAME', 'BQP_BIRTH_DATE', 'NLAD_ENROLLMENT_TYPE', 'ENROLLMENT_TYPE', 'CONSENT_FORM_AVAILABE', 'ID_PROOF_AVAILABE', 'DEVICE_REIMBURSEMENT__DATE', 'DEVICE_EXPECTED_RATE', 'REVIEWED_BY', 'SAME_MONTH_DISCONNECTION', 'ACTIVE_PERIOD']

def telgoo_sftp_connection_test():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect('52.2.57.121', port=22, username='maxsip', password='Ma(sI8cr^qS')
    print('connection made to telgoo')

    # Create an SFTP client
    sftp = ssh.open_sftp()
    directory_name = '/data/Activation Report'
    # Get the list of files in the directory with their attributes
    files = sftp.listdir_attr(directory_name)
    sorted_files = sorted(files, key=lambda x: x.st_mtime, reverse=True)
    print(sorted_files[0])
    

def declare_changes_dict() -> dict:
    changes = {}
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    columns = Merged_Customer_Databases.c.keys()
    for field in columns:
        changes[field] = {}
    return changes
         
def unavo_update_test_new():
    then = datetime.now()
    
    previous_day_file = '30-03-2024.csv'
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)

    changes = declare_changes_dict()

    unavo_to_merged_columns = column_mapping_unavo_to_merged('unavo_csv_to_merged')
    yesterday = time.mktime((datetime.today() - timedelta(days=1)).date().timetuple())
    customer_order_ids_with_changes = set()

    for unavo_col, merged_col in unavo_to_merged_columns.items():
        print(f'Time passed: {datetime.now()-then}. Starting {merged_col}')
        with engine.connect() as conn:
            stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', merged_col])\
                                            .where(Merged_Customer_Databases.c['Source_Database'] == 'Unavo')
            merged_db_contacts= conn.execute(stmt).all()
            stmt_selected_cols = stmt.selected_columns.keys()
        merged_db_unavo_contacts_with_col_value = {}
        for contact in merged_db_contacts:
            merged_db_unavo_contacts_with_col_value[contact[stmt_selected_cols.index('Customer_Order_ID')]] = contact[stmt_selected_cols.index(merged_col)]

        csv_unavo_contacts = {}
        with open(previous_day_file, encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile)
            fields = next(csvreader)
            order_id_index = fields.index('OrderId')
            current_selected_unavo_col_index = fields.index(unavo_col)
            for csv_reader_contact in csvreader:
                if len(csv_reader_contact) == len(fields):
                    csv_unavo_contacts[csv_reader_contact[order_id_index]] = csv_reader_contact[current_selected_unavo_col_index]

        for order_id, col_value in csv_unavo_contacts.items():
            if merged_db_unavo_contacts_with_col_value.get(order_id, None):

                if merged_col in ['Created_Date','Order_Date','Activation_Date','Deactivation_Date','DOB','Last_Used_Data',
                                    'Last_Used_Phone','ShipmentDate','deviceReimbursementDate','FollowupPeriod_CreatedDate',
                                    'BQP_BIRTH_DATE','DEVICE_REIMBURSEMENT__DATE',]:
                    if col_value:
                        col_value = parse(col_value).timestamp()
                if merged_db_unavo_contacts_with_col_value[order_id] != col_value:
                    customer_order_ids_with_changes.add(order_id)
                    changes.get(merged_col).append({
                        'Source_Database':'Unavo',
                        'Customer_Order_ID':order_id,
                        'Nlad_Subscriber_ID':'',
                        'Field_Name':merged_col,
                        'New_Value':col_value,
                        'Old_Value':merged_db_unavo_contacts_with_col_value[order_id],
                        'Change_Date':yesterday,
                        'Activation_Date':'',
                        'CLEC':'',
                        'Plan':'',
                        'Agent':'',
                        'Account_Status':'',
                        'ACP_Status':'',
                        'MDN':''
                    })
    print(f'Time taken to run whole thing{datetime.now() - then}')
    for k, v in changes.items():
        if len(v) > 0:
            print(f'{k}:{len(v)}')


    def declare_changes_dict():
        changes = {}
        engine = create_engine(path_to_db)
        metadata_obj = MetaData()
        metadata_obj.reflect(bind=engine)
        Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
        columns = Merged_Customer_Databases.c.keys()
        for field in columns:
            changes[field] = []
        return changes
   
def get_csv_unavo_contacts(csv_file:str, unavo_col:str) -> dict:
    csv_unavo_contacts = {}
    with open(csv_file, encoding='utf-8') as csvfile:
        csvreader = csv.reader(csvfile)
        fields = next(csvreader)
        order_id_index = fields.index('OrderId')
        current_selected_unavo_col_index = fields.index(unavo_col)
        for csv_reader_contact in csvreader:
            if len(csv_reader_contact) == len(fields):
                csv_unavo_contacts[csv_reader_contact[order_id_index]] = csv_reader_contact[current_selected_unavo_col_index]
    return csv_unavo_contacts

def get_merged_db_unavo_contacts_with_col_value(merged_col:str) -> dict:
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    with engine.connect() as conn:
        stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', merged_col])\
                                        .where(Merged_Customer_Databases.c['Source_Database'] == 'Unavo')
        merged_db_contacts= conn.execute(stmt).all()
        stmt_selected_cols = stmt.selected_columns.keys()

    merged_db_unavo_contacts_with_col_value = {}
    for contact in merged_db_contacts:
        merged_db_unavo_contacts_with_col_value[contact[stmt_selected_cols.index('Customer_Order_ID')]] = contact[stmt_selected_cols.index(merged_col)]
    
    return merged_db_unavo_contacts_with_col_value

def cleanup_changes_from_empty_dicts(changes:dict) -> dict:
    fields_to_delete = []
    for field, values in changes.items():
        if len(values) == 0:
            fields_to_delete.append(field)
    for field in fields_to_delete:
        del changes[field]
    return changes


def get_changes_and_customer_orders_for_insert_and_update(then, csv_file) -> tuple[dict, set, set]:
    changes = declare_changes_dict()

    unavo_to_merged_columns = column_mapping_unavo_to_merged('unavo_csv_to_merged')
    yesterday = time.mktime((datetime.today() - timedelta(days=1)).date().timetuple())
    customer_order_ids_with_changes = set()
    customer_order_ids_to_insert = set()

    for unavo_col, merged_col in unavo_to_merged_columns.items():
        print(f'Time passed: {datetime.now()-then}. Starting {merged_col}')
        
        merged_db_unavo_contacts_with_col_value = get_merged_db_unavo_contacts_with_col_value(merged_col)
        csv_unavo_contacts = get_csv_unavo_contacts(csv_file, unavo_col)

        for order_id, col_value in csv_unavo_contacts.items():
            if merged_db_unavo_contacts_with_col_value.get(order_id, None):
                if merged_col in ['Created_Date','Order_Date','Activation_Date','Deactivation_Date','DOB','Last_Used_Data',
                                    'Last_Used_Phone','ShipmentDate','deviceReimbursementDate','FollowupPeriod_CreatedDate',
                                    'BQP_BIRTH_DATE','DEVICE_REIMBURSEMENT__DATE']:
                    if col_value:
                        col_value = parse(col_value).timestamp()
                if merged_db_unavo_contacts_with_col_value[order_id] != col_value:
                    customer_order_ids_with_changes.add(order_id)
                    changes[merged_col][order_id] = {
                        'Source_Database':'Unavo',
                        'Customer_Order_ID':order_id,
                        'Nlad_Subscriber_ID':'',
                        'Field_Name':merged_col,
                        'New_Value':col_value,
                        'Old_Value':merged_db_unavo_contacts_with_col_value[order_id],
                        'Change_Date':yesterday,
                        'Activation_Date':'',
                        'CLEC':'',
                        'Plan':'',
                        'Agent':'',
                        'Account_Status':'',
                        'ACP_Status':'',
                        'MDN':''
                    }
            else:
                customer_order_ids_to_insert.add(order_id)
    changes = cleanup_changes_from_empty_dicts(changes)
    return changes, customer_order_ids_to_insert, customer_order_ids_with_changes

def get_fields_and_indexes_as_dict(fields):
    fields_dict = {}
    for field_num, field in enumerate(fields):
        fields_dict[field] = field_num
    return fields_dict


def column_mapping_unavo_to_merged(desired_map:str) -> dict:
    unavo_csv_to_merged = {
        'CLEC':'CLEC',
        'AgentName':'Agent',
        'Account':'Misc_Customer_ID',
        'OrderId':'Customer_Order_ID',
        'StatusType':'Account_Status',
        'FirstName':'First_Name',
        'LastName':'Last_Name',
        'SSN':'SSN',
        'Email':'Email',
        'MDN':'MDN', #10
        'MSID':'MSID',
        'State':'State',
        'City':'City',
        'ESN':'ESN',
        'CustomerPackage':'Plan',
        'CSA':'CSA',
        'StreetAddress':'Address1',
        'Apartment#orSuite#':'Address2',
        'Zipcode':'Zip',
        'PrimaryPhone':'Phone1', #20
        'DOB':'DOB',
        'ShipmentDate':'ShipmentDate',
        'Tracking#':'TrackingNum',
        'DeviceId':'DeviceId',
        'DataLastUsed':'Last_Used_Data',
        'VoiceTextLastUsed':'Last_Used_Phone',
        'NladProgram':'NladProgram',
        'LifelineSubscriberId':'LifelineSubscriberId',
        'LifelineStatus':'LifelineStatus',
        'NladErrorType':'NladErrorType', #30
        'IsAcp':'IsAcp',
        'AcpStatus':'ACP_Status',
        'deviceReimbursementDate':'deviceReimbursementDate',
        'DateAssigned':'Order_Date',
        'ActivationDate':'Activation_Date',
        'DeActivationDate':'Deactivation_Date',
        'PORTSTATUS':'PORTSTATUS',
        'IMEI':'IMEI',
        'IpAddress':'IpAddress',
        'AcpLifeLineCertificationType':'Qualifying_Program', #40
        'WirelessProviderType':'Service_Carrier',
        'IsProofOfBenefitsUploaded':'IsProofOfBenefitsUploaded',
        'IsIdentityProofUploaded':'IsIdentityProofUploaded',
        'IsTribal':'IsTribal',
        'TribalID':'TribalID',
        'National Verifier Application ID':'National_Verifier_Application_ID',
        'ACP SubscriberID':'NLAD_Subscriber_ID',
        'AgentFollowupPeriod':'AgentFollowupPeriod',
        'FollowupPeriod_CreatedDate':'FollowupPeriod_CreatedDate',
        'FollowupPeriodCreatedAuthor':'FollowupPeriodCreatedAuthor',#50
        'Model':'Device_Mfg',
        'ModelNumber':'Device_Model_Number',
        'DeviceType':'Device_Type',
        'FCC CheckBox':'FCC_CheckBox',
        'Software Consent':'Software_Consent',
        'Customer Price':'Customer_Price',
        'ConsentCheck':'ConsentCheck', #57
    }
    unavo_db_to_merged = {
        'CLEC':'CLEC',
        'AgentName':'Agent',
        'Account':'Misc_Customer_ID',
        'OrderId':'Customer_Order_ID',
        'StatusType':'Account_Status',
        'FirstName':'First_Name',
        'LastName':'Last_Name',
        'SSN':'SSN',
        'Email':'Email',
        'MDN':'MDN', #10
        'MSID':'MSID',
        'State':'State',
        'City':'City',
        'ESN':'ESN',
        'CustomerPackage':'Plan',
        'CSA':'CSA',
        'StreetAddress':'Address1',
        'ApartmentNumOrSuiteNum':'Address2',
        'Zipcode':'Zip',
        'PrimaryPhone':'Phone1', #20
        'DOB':'DOB',
        'ShipmentDate':'ShipmentDate',
        'TrackingNum':'TrackingNum',
        'DeviceId':'DeviceId',
        'DataLastUsed':'Last_Used_Data',
        'VoiceTextLastUsed':'Last_Used_Phone',
        'NladProgram':'NladProgram',
        'LifelineSubscriberId':'LifelineSubscriberId',
        'LifelineStatus':'LifelineStatus',
        'NladErrorType':'NladErrorType', #30
        'IsAcp':'IsAcp',
        'AcpStatus':'ACP_Status',
        'deviceReimbursementDate':'deviceReimbursementDate',
        'DateAssigned':'Order_Date',
        'ActivationDate':'Activation_Date',
        'DeActivationDate':'Deactivation_Date',
        'PORTSTATUS':'PORTSTATUS',
        'IMEI':'IMEI',
        'IpAddress':'IpAddress',
        'AcpLifeLineCertificationType':'Qualifying_Program', #40
        'WirelessProviderType':'Service_Carrier',
        'IsProofOfBenefitsUploaded':'IsProofOfBenefitsUploaded',
        'IsIdentityProofUploaded':'IsIdentityProofUploaded',
        'IsTribal':'IsTribal',
        'TribalID':'TribalID',
        'National_Verifier_Application_ID':'National_Verifier_Application_ID',
        'ACP_SubscriberID':'NLAD_Subscriber_ID',
        'AgentFollowupPeriod':'AgentFollowupPeriod',
        'FollowupPeriod_CreatedDate':'FollowupPeriod_CreatedDate',
        'FollowupPeriodCreatedAuthor':'FollowupPeriodCreatedAuthor',#50
        'Model':'Device_Mfg',
        'ModelNumber':'Device_Model_Number',
        'DeviceType':'Device_Type',
        'FCC_CheckBox':'FCC_CheckBox',
        'Software_Consent':'Software_Consent',
        'Customer_Price':'Customer_Price',
        'ConsentCheck':'ConsentCheck',
        'CreatedDate':'Created_Date' #58
    }

    if desired_map == 'unavo_db_to_merged':
        return unavo_db_to_merged
    elif desired_map == 'unavo_csv_to_merged':
        return unavo_csv_to_merged
    else:
        return 'Requested map doesnt exist, try "unavo_db_to_merged" or "unavo_csv_to_merged"'


def csv_to_db_upsert_test(path_to_db):
    csv_file = '30-03-2024.csv'
    then = datetime.now()

    changes, customer_order_ids_to_insert, customer_order_ids_with_changes = get_changes_and_customer_orders_for_insert_and_update(then, csv_file)

    print(f'Time passed: {datetime.now()-then}. Finished going through columns for changes')
    indexed_ids_for_insert = {k:k for k in customer_order_ids_to_insert}
    indexed_ids_for_update = {k:k for k in customer_order_ids_with_changes}
    whole_rows_for_update = []
    whole_rows_for_insert = []
    unavo_csv_to_merged_columns = column_mapping_unavo_to_merged('unavo_csv_to_merged')
    with open(csv_file, encoding='utf-8') as csvfile:
        csvreader = csv.reader(csvfile)
        fields = next(csvreader)
        fields_dict = get_fields_and_indexes_as_dict(fields)
        for contact_num, contact in enumerate(csvreader):
            if contact_num % 10000 == 0:
                print(contact_num)
            if len(contact) != len(fields):
                continue
            contact_order_id = contact[fields_dict['OrderId']]
            if indexed_ids_for_update.get(contact_order_id, None):
                whole_row_columns_and_data = {}
                for column, index in fields_dict.items():
                    if column in ['DateAssigned','ActivationDate','DeActivationDate','DOB','DataLastUsed',
                                    'VoiceTextLastUsed','ShipmentDate','deviceReimbursementDate','FollowupPeriod_CreatedDate']:
                        if contact[index]:
                            whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = parse(contact[index]).timestamp()
                    elif column == 'OrderId':
                        whole_row_columns_and_data['b_Customer_Order_ID'] = unavo_csv_to_merged_columns[column]
                    elif column == 'SSN':
                        if contact[index]:
                            temp_data = contact[index]
                            if len(temp_data) > 2:
                                if temp_data[-2] == '.':
                                    temp_data = temp_data[:-2]
                            digits = len(temp_data)
                            if digits == 4:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = temp_data
                            elif digits == 3:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'0{temp_data}'
                            elif digits == 2:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'00{temp_data}'
                            elif digits == 1:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'000{temp_data}'
                            else:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = temp_data
                        else:
                            whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = ''
                    else:
                        whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = contact[index]
                whole_rows_for_update.append(whole_row_columns_and_data)
                for contact_change_lists in changes.values():
                    if contact_change_lists.get(contact_order_id, None):
                        contact_change_dict = contact_change_lists[contact_order_id]
                        contact_change_dict['Activation_Date'] = contact[fields_dict['ActivationDate']]
                        contact_change_dict['Nlad_Subscriber_ID'] = contact[fields_dict['ACP SubscriberID']]
                        contact_change_dict['CLEC'] = contact[fields_dict['CLEC']]
                        contact_change_dict['Plan'] = contact[fields_dict['CustomerPackage']]
                        contact_change_dict['Agent'] = contact[fields_dict['AgentName']]
                        contact_change_dict['Account_Status'] = contact[fields_dict['StatusType']]
                        contact_change_dict['ACP_Status'] = contact[fields_dict['AcpStatus']]
                        contact_change_dict['MDN'] = contact[fields_dict['MDN']]
            elif indexed_ids_for_insert.get(contact_order_id, None):
                whole_row_columns_and_data = {}
                whole_row_columns_and_data['Source_Database'] = 'Unavo'
                whole_row_columns_and_data['Created_Date'] = yesterday
                for column, index in fields_dict.items():
                    if column in ['DateAssigned','ActivationDate','DeActivationDate','DOB','DataLastUsed',
                                    'VoiceTextLastUsed','ShipmentDate','deviceReimbursementDate','FollowupPeriod_CreatedDate']:
                        if contact[index]:
                            whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = parse(contact[index]).timestamp()
                    elif column == 'SSN':
                        if contact[index]:
                            temp_data = contact[index]
                            if len(temp_data) > 2:
                                if temp_data[-2] == '.':
                                    temp_data = temp_data[:-2]
                            digits = len(temp_data)
                            if digits == 4:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = temp_data
                            elif digits == 3:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'0{temp_data}'
                            elif digits == 2:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'00{temp_data}'
                            elif digits == 1:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = f'000{temp_data}'
                            else:
                                whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = temp_data
                        else:
                            whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = ''
                    else:
                        whole_row_columns_and_data[unavo_csv_to_merged_columns[column]] = contact[index]
                whole_rows_for_insert.append(whole_row_columns_and_data)
            
            length_of_rows_insert = len(whole_rows_for_insert)
            length_of_rows_update = len(whole_rows_for_update)
            if length_of_rows_insert % 500 == 0 and length_of_rows_insert > 0:
                # db_connect_and_insert(path_to_db, whole_rows_for_insert)
                whole_rows_for_insert = []
                print('500 rows submitted to db for inserting')
            if length_of_rows_update % 500 == 0 and length_of_rows_update > 0:
                # db_connect_and_update(path_to_db, whole_rows_for_update)
                whole_rows_for_update = []
                print('500 rows submitted to db for updating')
        if len(whole_rows_for_insert) > 0:
            # db_connect_and_insert(path_to_db, whole_rows_for_insert)
            whole_rows_for_insert = []
            print('Remainder of contacts for inserting have been submitted to db')
        print('DB inserting complete')
        if len(whole_rows_for_update) > 0:
            # db_connect_and_update(path_to_db, whole_rows_for_update)
            whole_rows_for_update = []
            print('Remainder of contacts for updating have been submitted to db')
        print('DB updating complete')
    
    for list_of_data_dicts in changes.values():
        engine = create_engine(path_to_db)
        metadata_obj = MetaData()
        metadata_obj.reflect(bind=engine)
        Merged_Database_Changes = Table("Merged_Database_Changes", metadata_obj, autoload_with=engine)
        # with engine.connect() as conn:
            # conn.execute(insert(Merged_Database_Changes), list_of_data_dicts)
            # conn.commit()
    print(f'Time passed: {datetime.now()-then}. Finished the whole function')






def change_analyze():
    changes = {
        'merged column': {
            'order id': {

            }
        }
    }

def get_customer_order_ids_to_insert(path_to_db:str):
    customer_order_ids_to_insert = set()
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    with engine.connect() as conn:
        db_order_ids = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID'])
                                        .where(Merged_Customer_Databases.c['Source_Database'] == 'Unavo')
                                        .limit(1)).all()
    db_order_ids_indexed = {k[0]:k[0] for k in db_order_ids}
    print(list(db_order_ids_indexed.keys())[0])


    
def column_mapping_telgoo_to_merged(desired_map:str) -> dict:
    telgoo_db_to_merged = {
        'CREATED_DATE':'Created_Date',
        'ORDER_DATE':'Order_Date',
        'ACTIVATION_DATE':'Activation_Date',
        'CUSTOMER_ID':'Misc_Customer_ID',
        'ENROLLMENT_ID':'Customer_Order_ID',
        'FIRST_NAME':'First_Name',
        'LAST_NAME':'Last_Name',
        'ADDRESS1':'Address1',
        'ADDRESS2':'Address2',
        'CITY':'City', #10
        'STATE':'State',
        'ZIP':'Zip',
        'SSN':'SSN',
        'DOB':'DOB',
        'IS_HOUSEHOLD':'IS_HOUSEHOLD',
        'IS_TRIBAL':'IS_TRIBAL',
        'IS_SHELTER':'IS_SHELTER',
        'IS_STATE_DB':'IS_STATE_DB',
        'ESN':'ESN',
        'MDN':'MDN', #20
        'EMAIL':'Email',
        'ALTERNATE_PHONE_NUMBER':'Phone1',
        'DEVICE_ID':'IMEI',
        'MAKE_MODEL':'Device_Model_Number',
        'NLAD_SUBSCRIBER_ID':'NLAD_Subscriber_ID',
        'STATE_DB_MATCHED':'STATE_DB_MATCHED',
        'RECORD_KEEPING':'RECORD_KEEPING',
        'ORDER_STATUS':'ORDER_STATUS',
        'PLAN':'Plan',
        'PRODUCT_TYPE':'Device_Type', #30
        'CARRIER':'Service_Carrier',
        'TABLET_SUBSIDY_QUALIFICATION':'TABLET_SUBSIDY_QUALIFICATION',
        'QUALIFYING_PROGRAM':'Qualifying_Program',
        'SOURCE':'SOURCE',
        'AUTHORIZE_CODE':'AUTHORIZE_CODE',
        'MASTER_AGENT_NAME':'MASTER_AGENT_NAME',
        'DISTRIBUTOR_AGENT_NAME':'DISTRIBUTOR_AGENT_NAME',
        'RETAILER_AGENT_NAME':'RETAILER_AGENT_NAME',
        'AGENT_NAME':'Agent',
        'DL_NUMBER':'DL_NUMBER', #40
        'WINBACK':'WINBACK',
        'ACCOUNT_STATUS':'Account_Status',
        'PAYMENT_TYPE':'PAYMENT_TYPE',
        'SPONSOR_ID':'SPONSOR_ID',
        'CURRENT_APS':'CURRENT_APS',
        'PORTIN_STATUS':'PORTIN_STATUS',
        'AGENT_LOGIN_ID':'Agent_LoginID',
        'DISCONNECT_REASON':'ACP_Status',
        'DEACTIVATION_DATE':'Deactivation_Date',
        'BQP_FIRST_NAME':'BQP_FIRST_NAME', #50
        'BQP_LAST_NAME':'BQP_LAST_NAME',
        'BQP_BIRTH_DATE':'BQP_BIRTH_DATE',
        'NLAD_ENROLLMENT_TYPE':'NLAD_ENROLLMENT_TYPE',
        'ENROLLMENT_TYPE':'ENROLLMENT_TYPE',
        'CONSENT_FORM_AVAILABE':'CONSENT_FORM_AVAILABE',
        'ID_PROOF_AVAILABE':'ID_PROOF_AVAILABE',
        'DEVICE_REIMBURSEMENT__DATE':'DEVICE_REIMBURSEMENT__DATE',
        'DEVICE_EXPECTED_RATE':'DEVICE_EXPECTED_RATE',
        'REVIEWED_BY':'REVIEWED_BY',
        'SAME_MONTH_DISCONNECTION':'SAME_MONTH_DISCONNECTION', #60
        'ACTIVE_PERIOD':'ACTIVE_PERIOD' #61
    }
    telgoo_csv_to_merged = {
        'CREATED DATE':'Created_Date',
        'ORDER DATE':'Order_Date',
        'ACTIVATION DATE':'Activation_Date',
        'CUSTOMER ID':'Misc_Customer_ID',
        'ENROLLMENT ID':'Customer_Order_ID',
        'FIRST NAME':'First_Name',
        'LAST NAME':'Last_Name',
        'ADDRESS1':'Address1',
        'ADDRESS2':'Address2',
        'CITY':'City', #10
        'STATE':'State',
        'ZIP':'Zip',
        'SSN':'SSN',
        'DOB':'DOB',
        'IS HOUSEHOLD':'IS_HOUSEHOLD',
        'IS TRIBAL':'IS_TRIBAL',
        'IS SHELTER':'IS_SHELTER',
        'IS STATE DB':'IS_STATE_DB',
        'ESN':'ESN',
        'MDN':'MDN', #20
        'EMAIL':'Email',
        'ALTERNATE PHONE NUMBER':'Phone1',
        'DEVICE ID':'IMEI',
        'MAKE MODEL':'Device_Model_Number',
        'NLAD SUBSCRIBER ID':'NLAD_Subscriber_ID',
        'STATE DB MATCHED':'STATE_DB_MATCHED',
        'RECORD KEEPING':'RECORD_KEEPING',
        'ORDER STATUS':'ORDER_STATUS',
        'PLAN':'Plan',
        'PRODUCT TYPE':'Device_Type', #30
        'CARRIER':'Service_Carrier',
        'TABLET SUBSIDY QUALIFICATION':'TABLET_SUBSIDY_QUALIFICATION',
        'QUALIFYING PROGRAM':'Qualifying_Program',
        'SOURCE':'SOURCE',
        'AUTHORIZE CODE':'AUTHORIZE_CODE',
        'MASTER AGENT NAME':'MASTER_AGENT_NAME',
        'DISTRIBUTOR AGENT NAME':'DISTRIBUTOR_AGENT_NAME',
        'RETAILER AGENT NAME':'RETAILER_AGENT_NAME',
        'AGENT NAME':'Agent',
        'DL NUMBER':'DL_NUMBER', #40
        'WINBACK':'WINBACK',
        'ACCOUNT STATUS':'Account_Status',
        'PAYMENT TYPE':'PAYMENT_TYPE',
        'SPONSOR ID':'SPONSOR_ID',
        'CURRENT APS':'CURRENT_APS',
        'PORTIN STATUS':'PORTIN_STATUS',
        'AGENT LOGIN ID':'Agent_LoginID',
        'DISCONNECT REASON':'ACP_Status',
        'DEACTIVATION DATE':'Deactivation_Date',
        'MEDICAID ID':'MEDICAID_ID', #50
        'BQP FIRST NAME':'BQP_FIRST_NAME',
        'BQP LAST NAME':'BQP_LAST_NAME',
        'BQP BIRTH DATE':'BQP_BIRTH_DATE',
        'NLAD ENROLLMENT TYPE':'NLAD_ENROLLMENT_TYPE',
        'ENROLLMENT TYPE':'ENROLLMENT_TYPE',
        'CONSENT FORM AVAILABE(Y/N)':'CONSENT_FORM_AVAILABE',
        'ID PROOF AVAILABE(Y/N)':'ID_PROOF_AVAILABE',
        'DEVICE REIMBURSEMENT  DATE':'DEVICE_REIMBURSEMENT__DATE',
        'DEVICE EXPECTED RATE':'DEVICE_EXPECTED_RATE',
        'REVIEWED BY':'REVIEWED_BY', #60
        'SAME MONTH DISCONNECTION':'SAME_MONTH_DISCONNECTION',
        'ACTIVE PERIOD':'ACTIVE_PERIOD' #62
    }
    
    if desired_map == 'telgoo_csv_to_merged':
        return telgoo_csv_to_merged
    elif desired_map == 'telgoo_db_to_merged':
        return telgoo_db_to_merged
    else:
        return 'Requested map doesnt exist, try "telgoo_csv_to_merged" or "telgoo_db_to_merged"'


def adding_telgoo_to_merged_db():
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    customerInfoTelgoo = Table("customerInfoTelgoo", metadata_obj, autoload_with=engine)
    fields = customerInfoTelgoo.c.keys()
    number_of_fields = len(fields)
    fields_dict = {}
    for field_num, field in enumerate(fields):
        fields_dict[field] = field_num

    telgoo_db_to_merged_column_map = column_mapping_telgoo_to_merged('telgoo_db_to_merged')
    with engine.connect() as conn:
        list_of_enroll_id_types = ['AMAX', 'WMAX', 'EMAX1', 'EMAX2', 'EMAX3', 'EMAX4', 'EMAX5', 'EMAX6', 'EMAX7', 'EMAX8', 'EMAX9']
        
        for type_of_enroll in list_of_enroll_id_types:
            print(type_of_enroll)
            telgoo_contacts_partial = conn.execute(select(customerInfoTelgoo)
                                        .where(customerInfoTelgoo.c['ENROLLMENT_ID'].like(f'{type_of_enroll}%'))).all()
            collection_of_row_data_for_submitting = []
            for contact in telgoo_contacts_partial:
                contact_row_data = [x for x in contact]
                if len(contact_row_data) == number_of_fields:
                    row_data_for_insert = {}
                    row_data_for_insert['Source_Database'] = 'Telgoo'
                    for field_name, field_index in fields_dict.items():
                        row_data_for_insert[telgoo_db_to_merged_column_map[field_name]] = contact_row_data[field_index]
                    collection_of_row_data_for_submitting.append(row_data_for_insert)
            
            stmt = insert(Merged_Customer_Databases)
            conn.execute(stmt, collection_of_row_data_for_submitting)
            conn.commit()
            print('SQL statement executed and committed')
                

def column_mapping_unavo_to_merged(desired_map:str) -> dict:
    unavo_csv_to_merged = {
        'CLEC':'CLEC',
        'AgentName':'Agent',
        'Account':'Misc_Customer_ID',
        'OrderId':'Customer_Order_ID',
        'StatusType':'Account_Status',
        'FirstName':'First_Name',
        'LastName':'Last_Name',
        'SSN':'SSN',
        'Email':'Email',
        'MDN':'MDN', #10
        'MSID':'MSID',
        'State':'State',
        'City':'City',
        'ESN':'ESN',
        'CustomerPackage':'Plan',
        'CSA':'CSA',
        'StreetAddress':'Address1',
        'Apartment#orSuite#':'Address2',
        'Zipcode':'Zip',
        'PrimaryPhone':'Phone1', #20
        'DOB':'DOB',
        'ShipmentDate':'ShipmentDate',
        'Tracking#':'TrackingNum',
        'DeviceId':'DeviceId',
        'DataLastUsed':'Last_Used_Data',
        'VoiceTextLastUsed':'Last_Used_Phone',
        'NladProgram':'NladProgram',
        'LifelineSubscriberId':'LifelineSubscriberId',
        'LifelineStatus':'LifelineStatus',
        'NladErrorType':'NladErrorType', #30
        'IsAcp':'IsAcp',
        'AcpStatus':'ACP_Status',
        'deviceReimbursementDate':'deviceReimbursementDate',
        'DateAssigned':'Order_Date',
        'ActivationDate':'Activation_Date',
        'DeActivationDate':'Deactivation_Date',
        'PORTSTATUS':'PORTSTATUS',
        'IMEI':'IMEI',
        'IpAddress':'IpAddress',
        'AcpLifeLineCertificationType':'Qualifying_Program', #40
        'WirelessProviderType':'Service_Carrier',
        'IsProofOfBenefitsUploaded':'IsProofOfBenefitsUploaded',
        'IsIdentityProofUploaded':'IsIdentityProofUploaded',
        'IsTribal':'IsTribal',
        'TribalID':'TribalID',
        'National Verifier Application ID':'National_Verifier_Application_ID',
        'ACP SubscriberID':'NLAD_Subscriber_ID',
        'AgentFollowupPeriod':'AgentFollowupPeriod',
        'FollowupPeriod_CreatedDate':'FollowupPeriod_CreatedDate',
        'FollowupPeriodCreatedAuthor':'FollowupPeriodCreatedAuthor',#50
        'Model':'Device_Mfg',
        'ModelNumber':'Device_Model_Number',
        'DeviceType':'Device_Type',
        'FCC CheckBox':'FCC_CheckBox',
        'Software Consent':'Software_Consent',
        'Customer Price':'Customer_Price',
        'ConsentCheck':'ConsentCheck', #57
    }
    unavo_db_to_merged = {
        'CLEC':'CLEC',
        'AgentName':'Agent',
        'Account':'Misc_Customer_ID',
        'OrderId':'Customer_Order_ID',
        'StatusType':'Account_Status',
        'FirstName':'First_Name',
        'LastName':'Last_Name',
        'SSN':'SSN',
        'Email':'Email',
        'MDN':'MDN', #10
        'MSID':'MSID',
        'State':'State',
        'City':'City',
        'ESN':'ESN',
        'CustomerPackage':'Plan',
        'CSA':'CSA',
        'StreetAddress':'Address1',
        'ApartmentNumOrSuiteNum':'Address2',
        'Zipcode':'Zip',
        'PrimaryPhone':'Phone1', #20
        'DOB':'DOB',
        'ShipmentDate':'ShipmentDate',
        'TrackingNum':'TrackingNum',
        'DeviceId':'DeviceId',
        'DataLastUsed':'Last_Used_Data',
        'VoiceTextLastUsed':'Last_Used_Phone',
        'NladProgram':'NladProgram',
        'LifelineSubscriberId':'LifelineSubscriberId',
        'LifelineStatus':'LifelineStatus',
        'NladErrorType':'NladErrorType', #30
        'IsAcp':'IsAcp',
        'AcpStatus':'ACP_Status',
        'deviceReimbursementDate':'deviceReimbursementDate',
        'DateAssigned':'Order_Date',
        'ActivationDate':'Activation_Date',
        'DeActivationDate':'Deactivation_Date',
        'PORTSTATUS':'PORTSTATUS',
        'IMEI':'IMEI',
        'IpAddress':'IpAddress',
        'AcpLifeLineCertificationType':'Qualifying_Program', #40
        'WirelessProviderType':'Service_Carrier',
        'IsProofOfBenefitsUploaded':'IsProofOfBenefitsUploaded',
        'IsIdentityProofUploaded':'IsIdentityProofUploaded',
        'IsTribal':'IsTribal',
        'TribalID':'TribalID',
        'National_Verifier_Application_ID':'National_Verifier_Application_ID',
        'ACP_SubscriberID':'NLAD_Subscriber_ID',
        'AgentFollowupPeriod':'AgentFollowupPeriod',
        'FollowupPeriod_CreatedDate':'FollowupPeriod_CreatedDate',
        'FollowupPeriodCreatedAuthor':'FollowupPeriodCreatedAuthor',#50
        'Model':'Device_Mfg',
        'ModelNumber':'Device_Model_Number',
        'DeviceType':'Device_Type',
        'FCC_CheckBox':'FCC_CheckBox',
        'Software_Consent':'Software_Consent',
        'Customer_Price':'Customer_Price',
        'ConsentCheck':'ConsentCheck',
        'CreatedDate':'Created_Date' #58
    }

    if desired_map == 'unavo_db_to_merged':
        return unavo_db_to_merged
    elif desired_map == 'unavo_csv_to_merged':
        return unavo_csv_to_merged
    else:
        return 'Requested map doesnt exist, try "unavo_db_to_merged" or "unavo_csv_to_merged"'




def adding_unavo_to_merged_db():
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    customerInfo = Table("customerInfo", metadata_obj, autoload_with=engine)

    fields = customerInfo.c.keys()
    number_of_fields = len(fields)
    fields_dict = {}
    for field_num, field in enumerate(fields):
        fields_dict[field] = field_num

    unavo_db_to_merged_column_map = column_mapping_unavo_to_merged('unavo_db_to_merged')
    with engine.connect() as conn:
        max_order_id = conn.scalar(select(func.max(customerInfo.c.OrderId)))
        if max_order_id == None:
            max_order_id = 0
        print(f'Max order id: {max_order_id}')
        starting_number = 0

        while starting_number < max_order_id:
            print(starting_number)
            unavo_contacts_partial = conn.execute(select(customerInfo)
                                                  .where(customerInfo.c.OrderId > starting_number,
                                                         customerInfo.c.OrderId <= starting_number + 500000))
            starting_number += 500000
            
            collection_of_row_data_for_submitting = []
            for contact in unavo_contacts_partial:
                contact_row_data = [x for x in contact]
                if len(contact_row_data) == number_of_fields:
                    row_data_for_insert = {}
                    row_data_for_insert['Source_Database'] = 'Unavo'
                    for field_name, field_index in fields_dict.items():
                        row_data_for_insert[unavo_db_to_merged_column_map[field_name]] = contact_row_data[field_index]
                    collection_of_row_data_for_submitting.append(row_data_for_insert)
            
            stmt = insert(Merged_Customer_Databases)
            conn.execute(stmt, collection_of_row_data_for_submitting)
            conn.commit()
            print('SQL statement executed and committed')
                
def writing_sample_data_to_excel():
    with engine.connect() as conn, Workbook('sample_data.xlsx') as wb:
        columns = Merged_Customer_Databases.c.keys()
        number_of_cols = len(columns)
        contacts = conn.execute(select(Merged_Customer_Databases).where(Merged_Customer_Databases.c['Source_Database'] == 'Unavo').limit(1000)).all()
        ws = wb.add_worksheet()
        ws.set_column(0, number_of_cols, 30)
        ws.write_row(0,0,columns)
        for row_num, contact in enumerate(contacts):
            ws.write_row(row_num + 1, 0, contact)

def telgoo_sftp_connection_test():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh.connect('52.2.57.121', port=22, username='maxsip', password='Ma(sI8cr^qS')

    sftp_client = ssh.open_sftp()
    directory_name = '/Terracom_Activation_Report'
    # directory_name = '/Maxsip_Activation_Report'
    
    files = sftp_client.listdir_attr(directory_name)
    sorted_files = sorted(files, key=lambda x: x.st_mtime, reverse=True)
    for file in sorted_files:
        if re.search('activation.*zip$', file.filename):
            print(file)
            # sftp_client.get(f'/Maxsip_Activation_Report/{file.filename}', f'/home/ubuntu/MaxsipReports/pythonFiles/{file.filename}')
            # with zipfile.ZipFile(file.filename, 'r') as zip_file:
            #     os.mkdir('zipfileholder')
            #     zip_file.extractall('zipfileholder')
            # os.remove(file.filename)
            break
    ssh.close()
    sftp_client.close()
    

def duplicate_audit():
    wb = load_workbook('Duplicate_Address_Audit.xlsx', read_only=True)
    ws = wb.active
    # library_of_sub_ids = {
    #     'sub_id':{
    #         'customer order id':{
    #             'Master Agent':'',
    #             'Agent':'',
    #             'Activation Date':'',
    #             'Deactivation Date':'',
    #             'Address':'',
    #             'City':'',
    #             'State':'',
    #             'Zip':''
    #         }
    #     }
    # }
    library_of_sub_ids = {}
    with engine.connect() as conn, Workbook('Duplicate audit output take 2.xlsx') as audit_wb:
        audit_ws = audit_wb.add_worksheet()
        audit_ws.set_column(0,14, 30)
        row_writer_counter = 1
        for row_num, row in enumerate(ws.rows):
            if row_num % 1000 == 0:
                print(row_num)
            if row_num == 0:
                fields = [x.value for x in row]
                fields_for_headers = [x for x in fields]
                fields_for_headers.insert(2, 'Customer Order ID')
                audit_ws.write_row(0,0,fields_for_headers)
                fields_dict = {}
                for field_num, field in enumerate(fields):
                    fields_dict[field] = field_num
                continue
            row_data = [x.value for x in row]
            row_sub_id = row_data[fields_dict['SUBSCRIBER ID']]

            if not library_of_sub_ids.get(row_sub_id, None):
                contacts_by_sub_id = conn.execute(
                    select(Merged_Customer_Databases.c['Customer_Order_ID', 'MASTER_AGENT_NAME', 'Agent', 'Activation_Date', 'Deactivation_Date', 'Address1', 'City', 'State', 'Zip'])
                    .where(Merged_Customer_Databases.c['NLAD_Subscriber_ID'] == row_sub_id)
                ).all()
                library_of_sub_ids[row_sub_id] = {}
                for contact in contacts_by_sub_id:
                    library_of_sub_ids[row_sub_id][contact[0]] = {
                        'Master Agent':contact[1],
                        'Agent':contact[2],
                        'Activation Date':contact[3],
                        'Deactivation Date':contact[4],
                        'Address':contact[5],
                        'City':contact[6],
                        'State':contact[7],
                        'Zip':contact[8]
                    }

            sub_id_dict = library_of_sub_ids.get(row_sub_id)
            for dict_num, (customer_order_id, value_dict) in enumerate(sub_id_dict.items()):
                row_data_to_write = [x for x in row_data]
                if value_dict['Deactivation Date']:
                    row_data_to_write[fields_dict['Date Deenrolled']] = datetime.fromtimestamp(value_dict['Deactivation Date']).strftime('%m/%d/%Y')
                else:
                    row_data_to_write[fields_dict['Date Deenrolled']] = ''
                if value_dict['Activation Date']:
                    row_data_to_write[fields_dict['date enrolled  ']] = datetime.fromtimestamp(value_dict['Activation Date']).strftime('%m/%d/%Y')
                else:
                    row_data_to_write[fields_dict['date enrolled  ']] = ''
                row_data_to_write[fields_dict['Master Agent ']] = value_dict['Master Agent']
                row_data_to_write[fields_dict['agent']] = value_dict['Agent']
                row_data_to_write[fields_dict['address']] = value_dict['Address']
                row_data_to_write[fields_dict['City']] = value_dict['City']
                row_data_to_write[fields_dict['State']] = value_dict['State']
                row_data_to_write[fields_dict['Zip']] = value_dict['Zip']
                row_data_to_write[fields_dict['Recovery Total ']] = ''
                row_data_to_write.insert(2, customer_order_id)
                audit_ws.write_row(row_writer_counter,0,row_data_to_write)
                row_writer_counter += 1



def add_changes_unavo_to_merged():
    list_of_insert_dicts = []
    with engine.connect() as conn:
        unavo_change_cols = unavoChanges.c.keys()
        fields_dict = {}
        for field_num, field in enumerate(unavo_change_cols):
            fields_dict[field] = field_num
        change_records = conn.execute(select(unavoChanges)).all()
        print('Changes loaded to memory')
        for record in change_records:
            list_of_insert_dicts.append({
                'Source_Database':'Unavo',
                'Customer_Order_ID':record[fields_dict['OrderId']],
                'Nlad_Subscriber_ID':record[fields_dict['Nlad_Subscriber_ID']],
                'Field_Name':record[fields_dict['Field_Name']],
                'New_Value':record[fields_dict['New_Value']],
                'Old_Value':record[fields_dict['Old_Value']],
                'Change_Date':record[fields_dict['Change_Date']],
                'Activation_Date':record[fields_dict['Activation_Date']],
                'CLEC':record[fields_dict['CLEC']],
                'Plan':record[fields_dict['Customer_Package']],
                'Agent': record[fields_dict['Agent_Name']],
                'Account_Status': record[fields_dict['Status_Type']],
                'ACP_Status': record[fields_dict['Acp_Status']],
                'MDN': record[fields_dict['MDN']]
            })
            if len(list_of_insert_dicts) == 500:
                stmt = insert(Merged_Database_Changes)
                conn.execute(stmt, list_of_insert_dicts)
                conn.commit()
                list_of_insert_dicts = []
        if len(list_of_insert_dicts) > 0:
            stmt = insert(Merged_Database_Changes)
            conn.execute(stmt, list_of_insert_dicts)
            conn.commit()

def add_changes_telgoo_to_merged():
    list_of_insert_dicts = []
    with engine.connect() as conn:
        telgoo_change_cols = telgooChanges.c.keys()
        fields_dict = {}
        for field_num, field in enumerate(telgoo_change_cols):
            fields_dict[field] = field_num
        change_records = conn.execute(select(telgooChanges)).all()
        print('Changes loaded to memory')
        for record in change_records:
            list_of_insert_dicts.append({
                'Source_Database':'Telgoo',
                'Customer_Order_ID':record[fields_dict['ENROLLMENT_ID']],
                'Nlad_Subscriber_ID':record[fields_dict['Nlad_Subscriber_Id']],
                'Field_Name':record[fields_dict['Field_Name']],
                'New_Value':record[fields_dict['New_Value']],
                'Old_Value':record[fields_dict['Old_Value']],
                'Change_Date':record[fields_dict['Change_Date']],
                'Plan':record[fields_dict['PLAN']],
                'Agent': record[fields_dict['AGENT_NAME']],
                'Account_Status': record[fields_dict['ACCOUNT_STATUS']],
                'ACP_Status': record[fields_dict['DISCONNECT_REASON']],
                'MDN': record[fields_dict['MDN']]
            })
            if len(list_of_insert_dicts) == 500:
                stmt = insert(Merged_Database_Changes)
                conn.execute(stmt, list_of_insert_dicts)
                conn.commit()
                list_of_insert_dicts = []
        if len(list_of_insert_dicts) > 0:
            stmt = insert(Merged_Database_Changes)
            conn.execute(stmt, list_of_insert_dicts)
            conn.commit()


def finding_changed_subs():
    old_sub_ids = ['W2WTFDH1E', '1FTFAT5X9', '12191257M']
    for file in os.listdir('/home/ubuntu/MaxsipReports/pythonFiles/zipfileholder'):
        with open(f'/home/ubuntu/MaxsipReports/pythonFiles/zipfileholder/{file}') as csvfile:
            csvreader = csv.reader(csvfile)
            fields = next(csvreader)
            sub_id_index = fields.index('NLAD SUBSCRIBER ID')
            customer_order_id = fields.index('ENROLLMENT ID')
            for contact in csvreader:
                if contact[sub_id_index] == 'WAW359017':
                    print(f'file name: {file}\nsub_id:{contact[sub_id_index]}\ncustomer order id: {contact[customer_order_id]}\n')


def get_from_merged_changes():
    with engine.connect() as conn, Workbook('change output.xlsx') as wb:
        ws = wb.add_worksheet()
        ws.set_column(0,10, 20)
        contacts = conn.execute(select(Merged_Database_Changes).where(Merged_Database_Changes.c['Field_Name'] == 'NLAD_Subscriber_ID')).all()
        columns = Merged_Database_Changes.c.keys()
        ws.write_row(0,0,columns)
        counter = 1
        for contact in contacts:
            row_data = [x for x in contact]
            ws.write_row(counter, 0, row_data)
            counter += 1



def updating_dates_in_changes():
    with engine.connect() as conn:
        contacts = conn.execute(select(Merged_Database_Changes.c['id', 'Activation_Date']).where(Merged_Database_Changes.c['Activation_Date'].like('__-%'))).all()
        list_for_update = []

        for contact in contacts:
            list_for_update.append({
                'b_id':contact[0],
                'Activation_Date':parse(contact[1]).timestamp()
            })
            if len(list_for_update) == 1000:
                conn.execute(update(Merged_Database_Changes)
                             .where(Merged_Database_Changes.c.id == bindparam('b_id'))
                             .values(Activation_Date = bindparam('Activation_Date'))
                             ,list_for_update)
                conn.commit()
                list_for_update = []

        if len(list_for_update) > 0:
            conn.execute(update(Merged_Database_Changes)
                             .where(Merged_Database_Changes.c.id == bindparam('b_id'))
                             .values(Activation_Date = bindparam('Activation_Date'))
                             ,list_for_update)
            conn.commit()

def double_check():
    with engine.connect() as conn:
        contacts = conn.execute(select(Merged_Database_Changes.c['id', 'Activation_Date']).where(Merged_Database_Changes.c['Activation_Date'].like('-%'))).all()
        print(len(contacts))


def finding_changed_subs():
    old_sub_ids = ['W2WTFDH1E', '1FTFAT5X9', '12191257M']
    with Workbook('old sub ids.xlsx') as wb:
        for file_num, file in enumerate(os.listdir('/home/ubuntu/MaxsipReports/pythonFiles/zipfileholder')):
            with open(f'/home/ubuntu/MaxsipReports/pythonFiles/zipfileholder/{file}') as csvfile:
                csvreader = csv.reader(csvfile)
                fields = next(csvreader)
                sub_id_index = fields.index('NLAD SUBSCRIBER ID')
                
                if file_num == 0:
                    ws = wb.add_worksheet()
                    ws.set_column(0,58,20)
                    ws.write_row(0,0,fields)
                    row_counter = 1
                for contact in csvreader:
                    if contact[sub_id_index] in old_sub_ids:
                        row_data = [x for x in contact]
                        ws.write_row(row_counter, 0, row_data)
                        row_counter += 1


def telgoo_sftp_connection_test():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh.connect('52.2.57.121', port=22, username='maxsip', password='Ma(sI8cr^qS')

    sftp_client = ssh.open_sftp()
    directory_name = '/Terracom_Activation_Report'
    # directory_name = '/Maxsip_Activation_Report'
    
    files = sftp_client.listdir_attr(directory_name)
    sorted_files = sorted(files, key=lambda x: x.st_mtime, reverse=True)
    booleanholder1 = False
    booleanholder2 = False
    for file in sorted_files:
        if re.search('activation.*zip$', file.filename):
            if booleanholder2 == True:
                print(file)
                sftp_client.get(f'/Terracom_Activation_Report/{file.filename}', f'/home/ubuntu/MaxsipReports/pythonFiles/{file.filename}')
                with zipfile.ZipFile(file.filename, 'r') as zip_file:
                    # os.mkdir('zipfileholder')
                    zip_file.extractall('zipfileholder')
                os.remove(file.filename)
                break
            if booleanholder1 == True:
                booleanholder2 = True
            booleanholder1 = True
    ssh.close()
    sftp_client.close()
    

def checking_contacts():
    order_id = 'EMAX10921933'
    with engine.connect() as conn:
        contact = conn.execute(select(customerInfoTelgoo).where(customerInfoTelgoo.c.ENROLLMENT_ID == order_id)).first()
        for k, v in zip(customerInfoTelgoo.c.keys(), contact):
            if k == 'DOB':
                print(f'{k}: {v}')
        # contact = conn.execute(select(Merged_Customer_Databases)
        #                        .where(Merged_Customer_Databases.c.Customer_Order_ID == order_id)).first()
        # for k, v in zip(Merged_Customer_Databases.c.keys(), contact):
        #     if k == 'DOB':
        #         print(f'{k}: {v}')

def double_check():
    with engine.connect() as conn:
        contacts = conn.execute(select(Merged_Database_Changes.c['id', 'Activation_Date']).where(Merged_Database_Changes.c['Activation_Date'].like('__-'))).all()
        print(len(contacts))


def telgoo_sftp_connection_test():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh.connect('52.2.57.121', port=22, username='maxsip', password='Ma(sI8cr^qS')

    sftp_client = ssh.open_sftp()
    # directory_name = '/Terracom_Activation_Report'
    directory_name = '/Maxsip_Activation_Report'
    
    files = sftp_client.listdir_attr(directory_name)
    sorted_files = sorted(files, key=lambda x: x.st_mtime, reverse=True)
    booleanholder1 = False
    booleanholder2 = False
    for file in sorted_files:
        if re.search('activation.*zip$', file.filename):
            print(file)
    ssh.close()
    sftp_client.close()
    

def checking_and_updating_field_name_changes():
    with engine.connect() as conn:
        contacts = conn.execute(select(distinct(Merged_Database_Changes.c['Field_Name']))).all()
        # contacts = conn.execute(select(Merged_Database_Changes).where(Merged_Database_Changes.c['Field_Name'] == None)).all()
        # print(len(contacts))
        for contact in contacts:
            print(contact)


def adding_groups_to_contacts_take_one():
    qualifications_to_be_a_duplicate = {
        0:['NLAD_Subscriber_ID'],
        1:['First_Name', 'Last_Name', 'Address1', 'Zip'],
        2:['DOB', 'SSN', 'First_Name'],
        3:['DOB', 'SSN', 'Last_Name'],
        4:['SSN', 'First_Name', 'Address1'],
        5:['SSN', 'Last_Name', 'Address1'],
        6:['DOB', 'First_Name', 'Address1'],
        7:['DOB', 'Last_Name', 'Address1']
    }

    stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', 'NLAD_Subscriber_ID', 'DOB', 
                                                'SSN', 'First_Name', 'Last_Name', 'Address1', 'Zip'])
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}

    with engine.connect() as conn:
        contacts = conn.execute(stmt).all()
    total_contacts = len(contacts)

    def check_for_falsy(contact:list) -> bool:
        for data_piece in contact:
            if bool(data_piece) == False:
                return False
        return True


    def get_matches_by_params(current_group:list, params:list) -> list:
        cust_order_ids_already_accounted_for = {c[fields_dict['Customer_Order_ID']] for c in current_group}
        unique_params = set()
        for contact in current_group:
            unique_params.add(tuple([contact[fields_dict[param]] for param in params]))
        print(f'Unique params: {unique_params}')
        matches = []
        for contact in contacts:
            order_id = contact[fields_dict['Customer_Order_ID']]
            if customer_order_id_to_group.get(order_id, None):
                continue
            contact_data = [x for x in contact]
            params_to_check = tuple([contact_data[fields_dict[param]] for param in params])

            if contact_data[fields_dict['Customer_Order_ID']] not in cust_order_ids_already_accounted_for and params_to_check in unique_params:
                print(f'Params to check: {params_to_check}')
                matches.append([x for x in contact_data])
        
        return matches

    def go_through_data_sets(group:list, index:int) -> list:
        complete_group = [x for x in group]
        for number, param_set in qualifications_to_be_a_duplicate.items():
            if number == index:
                break
            print(f'{number}:{param_set}')
            matches = get_matches_by_params(complete_group, param_set)
            for contact in matches:
                print(f'going through data again with new contact up to index {number}: {contact}')
                semi_group = go_through_data_sets([contact], number)
                complete_group += semi_group
        return complete_group

    group_number = 101

    customer_order_id_to_group = {}

    for contact in contacts:
        are_all_data_pieces_truthy = check_for_falsy(contact)
        if not are_all_data_pieces_truthy:
            customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = 1

    starting_length_of_cust_order_ids = len(customer_order_id_to_group.keys())

    start_time = datetime.now()
    for contact_num, contact in enumerate(contacts):
        row_data = [x for x in contact]
        order_id = row_data[fields_dict['Customer_Order_ID']]
        if customer_order_id_to_group.get(order_id, None):
            continue
        group = [row_data]
        complete_group = go_through_data_sets(group, index=8)
        for contact in complete_group:
            customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = group_number
        group_number += 1
        total_customer_order_ids_accounted_for = len(customer_order_id_to_group.keys()) - starting_length_of_cust_order_ids
        print(f'Percent completed: {(contact_num + 1)/total_contacts*100}')
        print(f'Estimated Time to finish: {((datetime.now() - start_time)/(total_customer_order_ids_accounted_for))*(total_contacts-total_customer_order_ids_accounted_for)}')
        print(f'Length of customer order id to group: {len(customer_order_id_to_group)}')

    with open("customer_order_id_to_group.json", "w") as outfile: 
        json.dump(customer_order_id_to_group, outfile)


def adding_groups_to_contacts_take_two():
    def check_for_falsy(contact:list) -> bool:
        for data_piece in contact:
            if bool(data_piece) == False:
                return False
        return True

    print(f'Starting available memory: {psutil.virtual_memory().available}')
    customer_order_id_to_group = {}

    then = datetime.now()
    qualifications_to_be_a_duplicate = {
        0:['NLAD_Subscriber_ID'],
        1:['First_Name', 'Last_Name', 'Address1', 'Zip'],
        2:['DOB', 'SSN', 'First_Name'],
        3:['DOB', 'SSN', 'Last_Name'],
        4:['SSN', 'First_Name', 'Address1'],
        5:['SSN', 'Last_Name', 'Address1'],
        6:['DOB', 'First_Name', 'Address1'],
        7:['DOB', 'Last_Name', 'Address1']
    }
    stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', 'NLAD_Subscriber_ID', 'DOB', 
                                                'SSN', 'First_Name', 'Last_Name', 'Address1', 'Zip'])
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}

    with engine.connect() as conn:
        contacts = conn.execute(stmt).all()
    total_contacts_length = len(contacts)
    for contact in contacts:
        are_all_data_pieces_truthy = check_for_falsy(contact)
        if not are_all_data_pieces_truthy:
            customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = 1

    starting_length_of_cust_order_ids = len(customer_order_id_to_group.keys())
    print(f'Number of contacts with missing data: {starting_length_of_cust_order_ids}')

    order_id_indexed_bank = {contact[fields_dict['Customer_Order_ID']]:contact for contact in contacts}

    total_bank = {}
    for params in qualifications_to_be_a_duplicate.values():
        total_bank['.'.join(params)] = {}
    for joined_params_names, param_values_to_order_id_list in total_bank.items():
        split_params = joined_params_names.split('.')
        print(f'starting: {split_params}')
        for contact in contacts:
            customer_order_id = contact[fields_dict['Customer_Order_ID']]
            if customer_order_id_to_group.get(customer_order_id, None):
                continue
            joined_param_values_of_contact = '.'.join([str(contact[fields_dict[param]]) for param in split_params])
            if param_values_to_order_id_list.get(joined_param_values_of_contact, None):
                param_values_to_order_id_list[joined_param_values_of_contact].append(customer_order_id)
            else:
                param_values_to_order_id_list[joined_param_values_of_contact] = [customer_order_id]
        print(f'Length of keys: {len(param_values_to_order_id_list.keys())}\nRemaining memory: {psutil.virtual_memory().available}')
    print(datetime.now() - then)

    group_number = 101


    def get_matches_by_params(current_group:list, params:list) -> list:
        matches = []
        for single_contact in current_group:
            # single_contact_order_id = single_contact[fields_dict['Customer_Order_ID']]
            joined_contact_params = '.'.join([str(single_contact[fields_dict[param]]) for param in params])
            list_of_common_order_ids = total_bank['.'.join(params)][joined_contact_params]
            for order_id_match in list_of_common_order_ids:
                if not customer_order_id_to_group.get(order_id_match, None):
                    matches.append(order_id_indexed_bank[order_id_match])
        return matches

    def go_through_data_sets(group:list, index:int) -> list:
        for contact_of_group in group:
            customer_order_id_to_group[contact_of_group[fields_dict['Customer_Order_ID']]] = group_number
        complete_group = [x for x in group]
        for number, param_set in qualifications_to_be_a_duplicate.items():
            if number == index:
                break
            print(f'{number}:{param_set}')
            matches = get_matches_by_params(complete_group, param_set)
            for contact_match in matches:
                print(f'going through data again with new contact up to index {number}: {contact}')
                semi_group = go_through_data_sets([contact_match], number)
                complete_group += semi_group
        return complete_group

    start_time = datetime.now()
    for contact_num, contact in enumerate(contacts):
        row_data = [x for x in contact]
        customer_order_id = row_data[fields_dict['Customer_Order_ID']]
        if customer_order_id_to_group.get(customer_order_id, None):
            continue
        group = [row_data]
        complete_group = go_through_data_sets(group, index=8)
        group_number += 1
        total_customer_order_ids_accounted_for = len(customer_order_id_to_group.keys()) - starting_length_of_cust_order_ids
        print(f'Percent completed: {(contact_num + 1)/total_contacts_length*100}')
        print(f'Estimated Time to finish: {((datetime.now() - start_time)/(total_customer_order_ids_accounted_for))*(total_contacts_length-total_customer_order_ids_accounted_for)}')
        print(f'Length of customer order id to group: {len(customer_order_id_to_group)}')
    print(f'Total function time: {datetime.now() - start_time}')
    with open("customer_order_id_to_group.json", "w") as outfile: 
        json.dump(customer_order_id_to_group, outfile)

def analyze_json():
    with open('customer_order_id_to_group.json') as f, engine.connect() as conn:
        data = json.load(f)
        list_for_update = []
        for customer_order_id, group_number in data.items():
            list_for_update.append({
                'b_Customer_Order_ID':customer_order_id,
                'Group_ID':group_number
            })
            if len(list_for_update) == 1000:
                conn.execute(update(Merged_Customer_Databases)
                             .where(Merged_Customer_Databases.c.Customer_Order_ID == bindparam('b_Customer_Order_ID'))
                             .values(Group_ID = bindparam('Group_ID'))
                             ,list_for_update)
                conn.commit()
                list_for_update = []

        if len(list_for_update) > 0:
            conn.execute(update(Merged_Customer_Databases)
                             .where(Merged_Customer_Databases.c.Customer_Order_ID == bindparam('b_Customer_Order_ID'))
                             .values(Group_ID = bindparam('Group_ID'))
                             ,list_for_update)
            conn.commit()


def adding_groups_to_contacts_take_three():
    print(f'Starting available memory: {psutil.virtual_memory().available}')
    customer_order_id_to_group = {}

    then = datetime.now()
    qualifications_to_be_a_duplicate = {
        0:['NLAD_Subscriber_ID'],
        1:['First_Name', 'Last_Name', 'Address1', 'Zip'],
        2:['DOB', 'SSN', 'First_Name'],
        3:['DOB', 'SSN', 'Last_Name'],
        4:['SSN', 'First_Name', 'Address1'],
        5:['SSN', 'Last_Name', 'Address1'],
        6:['DOB', 'First_Name', 'Address1'],
        7:['DOB', 'Last_Name', 'Address1']
    }
    stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', 'NLAD_Subscriber_ID', 'DOB', 
                                                'SSN', 'First_Name', 'Last_Name', 'Address1', 'Zip'])
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}

    with engine.connect() as conn:
        contacts = conn.execute(stmt).all()

    total_contacts_length = len(contacts)
    
    def check_for_falsy(contact:list) -> bool:
        for data_piece in contact:
            if bool(data_piece) == False:
                return False
        return True

    for contact in contacts:
        are_all_data_pieces_truthy = check_for_falsy(contact)
        if not are_all_data_pieces_truthy:
            customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = {
                'Group Number': 'Missing Data',
                'Match ID':'',
                'Matched Params':''
            }

    starting_length_of_cust_order_ids = len(customer_order_id_to_group.keys())
    print(f'Number of contacts with missing data: {starting_length_of_cust_order_ids}')

    order_id_indexed_bank = {contact[fields_dict['Customer_Order_ID']]:contact for contact in contacts}

    total_bank = {}
    for params in qualifications_to_be_a_duplicate.values():
        total_bank['.'.join(params)] = {}
    for joined_params_names, param_values_to_order_id_list in total_bank.items():
        split_params = joined_params_names.split('.')
        print(f'starting: {split_params}')
        for contact in contacts:
            customer_order_id = contact[fields_dict['Customer_Order_ID']]
            if customer_order_id_to_group.get(customer_order_id, None):
                continue
            joined_param_values_of_contact = '.'.join([str(contact[fields_dict[param]]) for param in split_params])
            if param_values_to_order_id_list.get(joined_param_values_of_contact, None):
                param_values_to_order_id_list[joined_param_values_of_contact].append(customer_order_id)
            else:
                param_values_to_order_id_list[joined_param_values_of_contact] = [customer_order_id]
        print(f'Length of keys: {len(param_values_to_order_id_list.keys())}\nRemaining memory: {psutil.virtual_memory().available}')
    print(datetime.now() - then)

    group_number = 101


    def get_matches_by_params(single_contact:list, params:list) -> list:
        matches = []
        joined_contact_params = '.'.join([str(single_contact[fields_dict[param]]) for param in params])
        list_of_common_order_ids = total_bank['.'.join(params)][joined_contact_params]
        for order_id_match in list_of_common_order_ids:
            if not customer_order_id_to_group.get(order_id_match, None):
                matches.append(order_id_indexed_bank[order_id_match])
        return matches

    def go_through_data_sets(contact_data:list, index:int, match_origin:str='Parent') -> list:
        if index == 8:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': group_number,
                'Match ID':match_origin,
                'Matched Params':''
            }
        else:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': group_number,
                'Match ID':match_origin,
                'Matched Params':index + 1
            }
        for number, param_set in qualifications_to_be_a_duplicate.items():
            if number == index:
                break
            matches = get_matches_by_params(contact_data, param_set)
            if matches:
                for contact_match in matches:
                    go_through_data_sets(contact_match, number, contact_data[fields_dict['Customer_Order_ID']])

    start_time = datetime.now()
    for contact_num, contact in enumerate(contacts):
        row_data = [x for x in contact]
        customer_order_id = row_data[fields_dict['Customer_Order_ID']]
        if customer_order_id_to_group.get(customer_order_id, None):
            continue
        go_through_data_sets(row_data, index=8)
        group_number += 1
        total_customer_order_ids_accounted_for = len(customer_order_id_to_group.keys()) - starting_length_of_cust_order_ids
        print(f'\rPercent completed: {(contact_num + 1)/total_contacts_length*100}, Estimated Time to finish: {((datetime.now() - start_time)/(total_customer_order_ids_accounted_for))*(total_contacts_length-total_customer_order_ids_accounted_for)}, Length of customer order id to group: {len(customer_order_id_to_group)}       ', end='', flush=True)
    print(f'\nTotal function time: {datetime.now() - start_time}')
    with open("take_two_customer_order_id_to_group.json", "w") as outfile: 
        json.dump(customer_order_id_to_group, outfile, indent=4)

def adding_groups_to_contacts_take_four():
    print(f'Starting available memory: {psutil.virtual_memory().available}')
    customer_order_id_to_group = {}

    then = datetime.now()
    qualifications_to_be_a_duplicate = {
        0:['NLAD_Subscriber_ID'],
        1:['First_Name', 'Last_Name', 'Address1', 'Zip'],
        2:['DOB', 'SSN', 'First_Name'],
        3:['DOB', 'SSN', 'Last_Name'],
        4:['SSN', 'First_Name', 'Address1'],
        5:['SSN', 'Last_Name', 'Address1'],
        6:['DOB', 'First_Name', 'Address1'],
        7:['DOB', 'Last_Name', 'Address1']
    }
    stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', 'NLAD_Subscriber_ID', 'DOB', 
                                                'SSN', 'First_Name', 'Last_Name', 'Address1', 'Zip'])
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}

    with engine.connect() as conn:
        contacts = conn.execute(stmt).all()

    total_contacts_length = len(contacts)
    
    def check_for_falsy(contact:list) -> bool:
        for data_piece in contact:
            if bool(data_piece) == False:
                return False
        return True

    for contact in contacts:
        are_all_data_pieces_truthy = check_for_falsy(contact)
        if not are_all_data_pieces_truthy:
            customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = {
                'Group Number': 'Missing Data',
                'Match ID':'',
                'Matched Params':''
            }

    starting_length_of_cust_order_ids = len(customer_order_id_to_group.keys())
    print(f'Number of contacts with missing data: {starting_length_of_cust_order_ids}')

    order_id_indexed_bank = {contact[fields_dict['Customer_Order_ID']]:contact for contact in contacts}

    total_bank = {}
    for params in qualifications_to_be_a_duplicate.values():
        total_bank['.'.join(params)] = {}
    for joined_params_names, param_values_to_order_id_list in total_bank.items():
        split_params = joined_params_names.split('.')
        print(f'starting: {split_params}')
        for contact in contacts:
            customer_order_id = contact[fields_dict['Customer_Order_ID']]
            if customer_order_id_to_group.get(customer_order_id, None):
                continue
            joined_param_values_of_contact = '.'.join([str(contact[fields_dict[param]]) for param in split_params])
            if param_values_to_order_id_list.get(joined_param_values_of_contact, None):
                param_values_to_order_id_list[joined_param_values_of_contact].append(customer_order_id)
            else:
                param_values_to_order_id_list[joined_param_values_of_contact] = [customer_order_id]
        print(f'Length of keys: {len(param_values_to_order_id_list.keys())}\nRemaining memory: {psutil.virtual_memory().available}')
    print(datetime.now() - then)

    group_number = 101


    def get_matches_by_params(total_group:list, params:list) -> list:
        matches = []
        for single_contact in total_group:
            joined_contact_params = '.'.join([str(single_contact[fields_dict[param]]) for param in params])
            list_of_common_order_ids = total_bank['.'.join(params)][joined_contact_params]
            for order_id_match in list_of_common_order_ids:
                if not customer_order_id_to_group.get(order_id_match, None):
                    matches.append(order_id_indexed_bank[order_id_match])
        return matches

    def go_through_data_sets(contact_data:list, index:int, match_origin:str='Parent'):
        total_group = [contact_data]
        if index == 8:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': group_number,
                'Match ID':match_origin,
                'Matched Params':''
            }
        else:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': group_number,
                'Match ID':match_origin,
                'Matched Params':index + 1
            }
        for number, param_set in qualifications_to_be_a_duplicate.items():
            if number == index:
                break
            matches = get_matches_by_params(total_group, param_set)
            if matches:
                total_group += matches
                for contact_match in matches:
                    go_through_data_sets(contact_match, number, contact_data[fields_dict['Customer_Order_ID']])

    start_time = datetime.now()
    for contact_num, contact in enumerate(contacts):
        row_data = [x for x in contact]
        customer_order_id = row_data[fields_dict['Customer_Order_ID']]
        if customer_order_id_to_group.get(customer_order_id, None):
            continue
        go_through_data_sets(row_data, index=8)
        group_number += 1
        total_customer_order_ids_accounted_for = len(customer_order_id_to_group.keys()) - starting_length_of_cust_order_ids
        print(f'\rPercent completed: {(contact_num + 1)/total_contacts_length*100}, Estimated Time to finish: {((datetime.now() - start_time)/(total_customer_order_ids_accounted_for))*(total_contacts_length-total_customer_order_ids_accounted_for)}, Length of customer order id to group: {len(customer_order_id_to_group)}       ', end='', flush=True)
    print(f'\nTotal function time: {datetime.now() - start_time}')
    # with open("take_three_customer_order_id_to_group.json", "w") as outfile: 
    #     json.dump(customer_order_id_to_group, outfile, indent=4)

    with engine.connect() as conn:
        list_for_update = []
        for customer_order_id, group_data in customer_order_id_to_group.items():
            list_for_update.append({
                'b_Customer_Order_ID':customer_order_id,
                'Group_ID':group_data['Group Number']
            })
            if len(list_for_update) == 1000:
                conn.execute(update(Merged_Customer_Databases)
                                .where(Merged_Customer_Databases.c.Customer_Order_ID == bindparam('b_Customer_Order_ID'))
                                .values(Group_ID = bindparam('Group_ID'))
                                ,list_for_update)
                conn.commit()
                list_for_update = []

        if len(list_for_update) > 0:
            conn.execute(update(Merged_Customer_Databases)
                                .where(Merged_Customer_Databases.c.Customer_Order_ID == bindparam('b_Customer_Order_ID'))
                                .values(Group_ID = bindparam('Group_ID'))
                                ,list_for_update)
            conn.commit()

def second_analyze_json():
    largest_group_numbers = set()
    counting_group_size = {}
    with open('take_three_customer_order_id_to_group.json') as f, Workbook('groups_over_ten.xlsx') as wb:
        ws = wb.add_worksheet()
        ws.set_column(0,4,15)
        ws.write_row(0,0,['Customer Order ID', 'Group Number', 'Match ID', 'Matched Parameter'])
        xl_counter = 1
        data = json.load(f)
        for order_id, contact_data in data.items():
            if issubclass(int, type(contact_data['Group Number'])):
                if counting_group_size.get(contact_data['Group Number']):
                    counting_group_size[contact_data['Group Number']] += 1
                else:
                    counting_group_size[contact_data['Group Number']] = 1
        print('Finished part one')
        for group_num, qty in counting_group_size.items():
            if qty >= 20:
                print(f'{group_num}: {qty}')
                largest_group_numbers.add(group_num)
        print('Finished part two')
        for order_id, contact_data in data.items():
            if issubclass(int, type(contact_data['Group Number'])):
                if contact_data['Group Number'] in largest_group_numbers:
                    ws.write_row(xl_counter, 0, [order_id, contact_data['Group Number'], contact_data['Match ID'], contact_data['Matched Params']])
                    xl_counter += 1

def add_to_excel():
    holding_data = {}
    readwb = load_workbook('groups_over_ten.xlsx')
    readws = readwb.active
    for row_num, row in enumerate(readws.rows):
        if row_num == 0:
            fields = [x.value for x in row]
            continue
        row_data = [x.value for x in row]
        holding_data[row_data[0]] = row_data

    with engine.connect() as conn:
        order_ids_collection = []
        for order_id in holding_data.keys():
            order_ids_collection.append(order_id)
            if len(order_ids_collection) == 900:
                db_contacts_data = conn.execute(
                    select(Merged_Customer_Databases.c['Customer_Order_ID', 'Activation_Date', 'Account_Status', 
                                                       'Deactivation_Date', 'NLAD_Subscriber_ID', 'First_Name', 
                                                       'Last_Name', 'Address1', 'Zip', 'DOB', 'SSN', 'Source_Database', 'MASTER_AGENT_NAME', 'Agent'])
                    .where(Merged_Customer_Databases.c['Customer_Order_ID'].in_(order_ids_collection))
                ).all()
                for db_contact in db_contacts_data:
                    row_data = [x for x in db_contact]
                    if row_data[1]:
                        row_data[1] = datetime.fromtimestamp(float(row_data[1])).strftime('%Y-%m-%d')
                    if row_data[3]:
                        row_data[3] = datetime.fromtimestamp(float(row_data[3])).strftime('%Y-%m-%d')
                    if row_data[9]:
                        row_data[9] = datetime.fromtimestamp(float(row_data[9])).strftime('%Y-%m-%d')
                    holding_data[row_data[0]] += row_data[1:]
                order_ids_collection = []
        if order_ids_collection:
            db_contacts_data = conn.execute(
                select(Merged_Customer_Databases.c['Customer_Order_ID', 'Activation_Date', 'Account_Status', 
                                                    'Deactivation_Date', 'NLAD_Subscriber_ID', 'First_Name', 
                                                    'Last_Name', 'Address1', 'Zip', 'DOB', 'SSN', 'Source_Database'])
                .where(Merged_Customer_Databases.c['Customer_Order_ID'].in_(order_ids_collection))
            ).all()
            for db_contact in db_contacts_data:
                row_data = [x for x in db_contact]
                if row_data[1]:
                    row_data[1] = datetime.fromtimestamp(float(row_data[1])).strftime('%Y-%m-%d')
                if row_data[9]:
                    row_data[9] = datetime.fromtimestamp(float(row_data[9])).strftime('%Y-%m-%d')
                holding_data[row_data[0]] += row_data[1:]

    with Workbook('output.xlsx') as wb:
        ws = wb.add_worksheet()
        ws.set_column(0,18,17)
        ws.write_row(0,0,['Customer Order Id', 'Group Number', 'Match ID', 'Matched Parameter', 
                          'Activation Date', 'Account Status', 'Deactivation Date', 'Subscriber ID', 
                          'First Name', 'Last Name', 'Address1', 'Zip', 'DOB', 'SSN', 'Source Database', 'MASTER_AGENT_NAME', 'Agent'])
        excel_counter = 1
        for writing_list in holding_data.values():
            ws.write_row(excel_counter, 0, writing_list)
            excel_counter += 1

def get_largest_group_number():
    largest_group_num = 0
    groups_min_two_members = 0
    counting_group_size = {}
    with open('take_two_customer_order_id_to_group.json') as jsonfile:
        data = json.load(jsonfile)
    for value_dict in data.values():
        if issubclass(int, type(value_dict['Group Number'])):
            if counting_group_size.get(value_dict['Group Number']):
                counting_group_size[value_dict['Group Number']] += 1
            else:
                counting_group_size[value_dict['Group Number']] = 1
            if value_dict['Group Number'] > largest_group_num:
                largest_group_num = value_dict['Group Number']
    print(largest_group_num)
    for v in counting_group_size.values():
        if v > 1:
            groups_min_two_members += 1
    print(groups_min_two_members)


def activation_groups_with_agents():
    start_date = '03-23-2024'
    timestamp_start_date = parse(start_date).timestamp()
    end_date = '03-24-2024'
    timestamp_end_date = parse(end_date).timestamp()
    with engine.connect() as conn:
        stmt = select(Merged_Customer_Databases.c['Source_Database', 'MASTER_AGENT_NAME', 'Agent', 'Group_ID']).where(
            Merged_Customer_Databases.c['Activation_Date'] >= timestamp_start_date,
            Merged_Customer_Databases.c['Activation_Date'] <= timestamp_end_date,
        )
        contacts = conn.execute(stmt).all()
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}

    def get_group_sizes(contacts:list) -> dict:
        result = {}
        for contact in contacts:
            if result.get(contact[fields_dict['Group_ID']], None):
                result[contact[fields_dict['Group_ID']]] += 1
            else:
                result[contact[fields_dict['Group_ID']]] = 1
        return result
    
    group_to_size = get_group_sizes(contacts)

    def get_qty_range(number:int) -> str:
        if number == 1:
            return '1'
        elif 5 >= number >= 2:
            return '2-5'
        elif 10 >= number >= 6:
            return '6-10'
        else:
            return '10+'

    def get_master_agent_quantity(contacts:list) -> list[list]:
        result = []
        master_agents_and_agents = {}
        for contact in contacts:
            contact_master_agent_and_agent = '.'.join([contact[fields_dict['MASTER_AGENT_NAME']], contact[fields_dict['Agent']]])
            if master_agents_and_agents.get(contact_master_agent_and_agent, None):
                master_agents_and_agents[contact_master_agent_and_agent][contact[fields_dict['Source_Database']]] += 1
                master_agents_and_agents[contact_master_agent_and_agent]['Total'] += 1
                master_agents_and_agents[contact_master_agent_and_agent][get_qty_range(group_to_size[contact[fields_dict['Group_ID']]])] += 1
            else:
                master_agents_and_agents[contact_master_agent_and_agent] = {
                    'Total':0,
                    'Terracom':0,
                    'Unavo':0,
                    'Telgoo':0,
                    '1':0,
                    '2-5':0,
                    '6-10':0,
                    '10+':0
                }
                master_agents_and_agents[contact_master_agent_and_agent][contact[fields_dict['Source_Database']]] += 1
                master_agents_and_agents[contact_master_agent_and_agent]['Total'] += 1
                master_agents_and_agents[contact_master_agent_and_agent][get_qty_range(group_to_size[contact[fields_dict['Group_ID']]])] += 1
        for master_agent_agent, data_dict in master_agents_and_agents.items():
            master_agent_agent_split = master_agent_agent.split('.')
            data_list = [x for x in data_dict.values()]
            result.append(master_agent_agent_split + data_list)
        result.sort(key= lambda x:x[2], reverse=True)
        return result

    master_agent_to_qty = get_master_agent_quantity(contacts)

    for masteragentlist in master_agent_to_qty[:10]:
        print(masteragentlist)


def playing_with_os():
    for k, v in os.environ.items():
        print(f'{k}: {v}')


def get_files_from_sftp():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh.connect('52.2.57.121', port=22, username='maxsip', password='Ma(sI8cr^qS')

    sftp_client = ssh.open_sftp()
    remote_directory = '/Maxsip_Activation_Report'
    local_directory = '/home/ubuntu/MaxsipReports/pythonFiles'
    
    files = sftp_client.listdir_attr(remote_directory)
    sorted_files = sorted(files, key=lambda x: x.st_mtime, reverse=True)
    # print(sorted_files[1])
    for file in sorted_files:
        if re.search('activation.*zip$', file.filename):
            sftp_client.get(f'/Maxsip_Activation_Report/{file.filename}', f'{local_directory}/{file.filename}')
            with zipfile.ZipFile(file.filename, 'r') as zip_file:
                os.mkdir('/home/ubuntu/MaxsipReports/pythonFiles/maxsip_zip_file_holder')
                zip_file.extractall('maxsip_zip_file_holder')
            os.remove(file.filename)
            break
    ssh.close()
    sftp_client.close()
    return '/home/ubuntu/MaxsipReports/pythonFiles/maxsip_zip_file_holder'

def checking_column_data(current_file):
    telgoo_list_of_column_names = ['#','CREATED DATE', 'ORDER DATE', 'ACTIVATION DATE', 'CUSTOMER ID', 'ENROLLMENT ID', 'FIRST NAME', 
                                'LAST NAME', 'ADDRESS1', 'ADDRESS2', 'CITY', 'STATE', 'ZIP', 'SSN', 
                                'DOB', 'IS HOUSEHOLD', 'IS TRIBAL', 'IS SHELTER', 'IS STATE DB', 'ESN', 'MDN', 
                                'EMAIL', 'ALTERNATE PHONE NUMBER', 'DEVICE ID', 'MAKE MODEL', 'NLAD SUBSCRIBER ID', 'STATE DB MATCHED', 'RECORD KEEPING', 
                                'ORDER STATUS', 'PLAN', 'PRODUCT TYPE', 'CARRIER', 'TABLET SUBSIDY QUALIFICATION', 'QUALIFYING PROGRAM', 'SOURCE', 
                                'AUTHORIZE CODE', 'MASTER AGENT NAME', 'DISTRIBUTOR AGENT NAME', 'RETAILER AGENT NAME', 'AGENT NAME', 'DL NUMBER', 'WINBACK', 
                                'ACCOUNT STATUS', 'PAYMENT TYPE', 'SPONSOR ID', 'CURRENT APS', 'PORTIN STATUS', 'AGENT LOGIN ID', 'DISCONNECT REASON', 
                                'DEACTIVATION DATE', 'MEDICAID ID', 'BQP FIRST NAME', 'BQP LAST NAME', 'BQP BIRTH DATE', 'NLAD ENROLLMENT TYPE', 'ENROLLMENT TYPE', 
                                'CONSENT FORM AVAILABE(Y/N)', 'ID PROOF AVAILABE(Y/N)', 'DEVICE REIMBURSEMENT  DATE', 'DEVICE EXPECTED RATE', 'REVIEWED BY', 
                                'SAME MONTH DISCONNECTION', 'ACTIVE PERIOD']
    with open(current_file) as csvfile:
        csvreader = csv.reader(csvfile)
        fields = next(csvreader)
    if telgoo_list_of_column_names == fields:
        return True
    else:
        for field in fields:
            if field not in telgoo_list_of_column_names:
                print(field)
        print('\n')
        for field in telgoo_list_of_column_names:
            if field not in fields:
                print(field)
        return False

def vacuum():
    stmt = text('VACUUM')
    with engine.connect() as conn:
        conn.execute(stmt)
        
def check_group_numbers():
    with engine.connect() as conn:
        contacts = conn.execute(select(func.count()).select_from(Merged_Customer_Databases).where(Merged_Customer_Databases.c['Account_Status'] == 'Active')).first()
        # contacts = conn.execute(select(distinct(Merged_Customer_Databases.c['Account_Status']))).all()
        print(contacts)

def check_changes():
    with engine.connect() as conn:
        changes = conn.execute(select(func.count()).select_from(Merged_Database_Changes).where(Merged_Database_Changes.c['Field_Name'] == 'Activation_Date')).first()
        changes = conn.execute(select(distinct(Merged_Database_Changes.c['Field_Name']))).all()
        for change in changes:
            if change[0] == 'Activation_Date':
                print(change[0])
    # print(changes)

def check_distinct_dates():
    adjustment_from_to_map = {
        "DEVICE_ID":"IMEI",
        "MAKE_MODEL":"Device_Model_Number",
        "PRODUCT_TYPE":"Device_Type",
        "DISCONNECT_REASON":"ACP_Status",
        "NEW_CONTACTS":"New Contacts",
        "New_Unavo_Records":"New Contacts",
        "StatusType":"Account_Status",
        "AcpStatus":"ACP_Status",
        "Model":"Device_Mfg",
        "ModelNumber":"Device_Model_Number",
        "DeviceType":"Device_Type"
    }
    with engine.connect() as conn, Workbook('accounts.xlsx') as wb:
        ws = wb.add_worksheet()
        contacts = conn.execute(select(Merged_Database_Changes).where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status')).all()
        ws.write_row(0,0,list(Merged_Database_Changes.c.keys()))
        counter = 1
        for contact in contacts:
            row_data = [x for x in contact]
            if row_data[7]:
                row_data[7] = datetime.fromtimestamp(row_data[7]).strftime('%m-%d-%Y %H:%M:%S')
            ws.write_row(counter, 0, row_data)
            counter += 1
        # contacts = conn.execute(select(Merged_Database_Changes.c['Change_Date'])
        #                         .where(Merged_Database_Changes.c['Field_Name'] == 'IMEI')
        #                         .order_by(Merged_Database_Changes.c['Change_Date'].asc()).limit(1)).first()[0]
        # date = datetime.fromtimestamp(contacts)
        # print(date)

        # x = conn.execute(select(distinct(Merged_Database_Changes.c['Field_Name']))).all()
        # x = [y[0] for y in x]
        # db_fields = list(Merged_Customer_Databases.c.keys())
        # for t in x:
        #     if not t in db_fields:
        #         print(t)

        # for key, val in adjustment_from_to_map.items():
        #     contacts = conn.execute(select(Merged_Database_Changes.c['id']).where(Merged_Database_Changes.c['Field_Name'] == key)).all()
        #     list_of_dicts = []
        #     for contact in contacts:
        #         list_of_dicts.append({
        #             'b_id': contact[0],
        #             'Field_Name':val
        #         })
        #     conn.execute(update(Merged_Database_Changes)
        #                  .where(Merged_Database_Changes.c['id'] == bindparam('b_id'))
        #                  .values(Field_Name = bindparam('Field_Name'))
        #                  , list_of_dicts)
        #     conn.commit()

        # dates = conn.execute(select(distinct(Merged_Database_Changes.c['Change_Date']))
        #                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status')
        #                      .order_by(Merged_Database_Changes.c['Change_Date'])).all()
        # for date in dates:
        #     if date[0]:
        #         print(datetime.fromtimestamp(date[0]))



def populate_active_customer_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status'])).all()
        main_table_contacts = {contact[0]:contact[1] for contact in main_table_contacts}
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            as_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in as_changes:
                main_table_contacts[contact[0]] = contact[1]
            counter = 0
            for val in main_table_contacts.values():
                if val == 'Active':
                    counter += 1
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Active Customer Count',
                             count=counter
                         ))
            conn.commit()
            print(f'Total Active Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {counter}')
            # print(f'Total contacts for day: {datetime.fromtimestamp(target_timestamp).date()} = {len(main_table_contacts)}')

def populate_total_customer_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '')).all()
        main_table_contacts = {contact[0]:contact[1] for contact in main_table_contacts}
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
                    
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Total Customer Count',
                             count=len(main_table_contacts)
                         ))
            conn.commit()
            print(f'Total Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {len(main_table_contacts)}')

def populate_thirty_day_active_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status', 'Activation_Date'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '')).all()
        main_table_contacts = {
            contact[0]:{
                'Account_Status':contact[1], 
                'Activation_Date':contact[2]
            } for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            month_before_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta + 29)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            as_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in as_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['Account_Status'] = contact[1]
            counter = 0
            for val in main_table_contacts.values():
                if val['Account_Status'] == 'Active' and val['Activation_Date'] < month_before_timestamp:
                    counter += 1
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Thirty Day Customer Count',
                             count=counter
                         ))
            conn.commit()
            print(f'Total Active Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {counter}')



# Status Report populate scripts
def populate_active_customer_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != '')).all()
        main_table_contacts = {contact[0]:contact[1] for contact in main_table_contacts}
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,37):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            as_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in as_changes:
                main_table_contacts[contact[0]] = contact[1]
            counter = 0
            for val in main_table_contacts.values():
                if val == 'Active':
                    counter += 1
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Active Customer Count',
                             count=counter
                         ))
            conn.commit()
            print(f'Total Active Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {counter}')
            # print(f'Total contacts for day: {datetime.fromtimestamp(target_timestamp).date()} = {len(main_table_contacts)}')

def populate_total_customer_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '')).all()
        main_table_contacts = {contact[0]:contact[1] for contact in main_table_contacts}
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
                    
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Total Customer Count',
                             count=len(main_table_contacts)
                         ))
            conn.commit()
            print(f'Total Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {len(main_table_contacts)}')

def populate_thirty_day_active_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Account_Status', 'Activation_Date'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != '')).all()
        main_table_contacts = {
            contact[0]:{
                'Account_Status':contact[1], 
                'Activation_Date':contact[2]
            } for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,37):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            month_before_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta + 30)).date().timetuple())
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            as_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in as_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['Account_Status'] = contact[1]
            counter = 0
            for val in main_table_contacts.values():
                if val['Account_Status'] == 'Active' and val['Activation_Date'] <= month_before_timestamp:
                    counter += 1
            conn.execute(insert(Status_Report_Counts)
                         .values(
                             date=target_timestamp,
                             field='Thirty Day Customer Count',
                             count=counter
                         ))
            conn.commit()
            print(f'Total 30+ Active Customers for day: {datetime.fromtimestamp(target_timestamp).date()} = {counter}')


# Status Report Scripts second half
def populate_enrollment_types():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'ENROLLMENT_TYPE'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['Account_Status'] == 'Active',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != '',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != None)).all()
        main_table_contacts = {
            contact[0]:contact[1] for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            print(f"Starting day: {datetime.fromtimestamp(target_timestamp).strftime('%m/%d/%Y')}")
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            result = {}
            list_of_inserts = []
            for e_type in main_table_contacts.values():
                if result.get(e_type, None):
                    result[e_type] += 1
                else:
                    result[e_type] = 1
            for e_type, enroll_count in result.items():
                list_of_inserts.append({
                    'date':target_timestamp,
                    'field':e_type,
                    'count':enroll_count
                })


            stmt = insert(Status_Report_Counts)
            conn.execute(stmt, list_of_inserts)
            conn.commit()

def populate_enrollment_types_take_two():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'ENROLLMENT_TYPE', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != '')).all()
        main_table_contacts = {
            contact[0]:{
                'enroll_type':contact[1],
                'acct_status':contact[2]
            } for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,37):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            print(f"Starting day: {datetime.fromtimestamp(target_timestamp).strftime('%m/%d/%Y')}")
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            a_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in a_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['acct_status'] = contact[1]

            result = {}
            for e_type in main_table_contacts.values():
                if result.get(e_type['enroll_type'], None):
                    if e_type['acct_status'] == 'Active':
                        result[e_type['enroll_type']] += 1
                else:
                    result[e_type['enroll_type']] = 1

            list_of_inserts = []
            for e_type, enroll_count in result.items():
                list_of_inserts.append({
                    'date':target_timestamp,
                    'field':e_type,
                    'count':enroll_count
                })


            stmt = insert(Status_Report_Counts)
            conn.execute(stmt, list_of_inserts)
            conn.commit()


# Master Agent and Agent populate scripts
def master_agent_and_agent_active_customer():
    with engine.connect() as conn:
        agents = conn.execute(select(Merged_Customer_Databases.c['Agent'])
                              .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                     Merged_Customer_Databases.c['Account_Status'] == 'Active',
                                     Merged_Customer_Databases.c['Agent'] != '')).all()
    result = {}
    for agent in agents:
        if result.get(agent[0], None):
            result[agent[0]] += 1
        else:
            result[agent[0]] = 1
    result = {k: v for k, v in sorted(result.items(), key=lambda item: item[1], reverse=True)}
    for num,(k, v) in enumerate(result.items()):
        print(f'{k}: {v}')
        if num == 5:
            break


def populate_agent_active_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'Agent', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['Agent'] != '')).all()
        main_table_contacts = {
            contact[0]:{
                'agent':contact[1],
                'acct_status':contact[2]
            } for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            print(f"Starting day: {datetime.fromtimestamp(target_timestamp).strftime('%m/%d/%Y')}")
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            agent_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Agent',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in agent_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['agent'] = contact[1]

            account_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in account_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['acct_status'] = contact[1]

            result = {}
            list_of_inserts = []
            for agent in main_table_contacts.values():
                if result.get(agent['agent'], None):
                    if agent['acct_status'] == 'Active':
                        result[agent['agent']] += 1
                else:
                    result[agent['agent']] = 1
            for agent, enroll_count in result.items():
                list_of_inserts.append({
                    'date':target_timestamp,
                    'name':agent,
                    'count':enroll_count,
                    'type':'Active Agent'
                })


            stmt = insert(Master_Agent_And_Agent_Active_Counts)
            conn.execute(stmt, list_of_inserts)
            conn.commit()



def populate_master_agent_active_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'MASTER_AGENT_NAME', 'Account_Status'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['MASTER_AGENT_NAME'] != '')).all()
        main_table_contacts = {
            contact[0]:{
                'master_agent':contact[1],
                'acct_status':contact[2]
            } for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            print(f"Starting day: {datetime.fromtimestamp(target_timestamp).strftime('%m/%d/%Y')}")
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            agent_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'MASTER_AGENT_NAME',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in agent_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['agent'] = contact[1]

            account_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'Account_Status',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in account_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]]['acct_status'] = contact[1]

            result = {}
            list_of_inserts = []
            for master_agent in main_table_contacts.values():
                if result.get(master_agent['master_agent'], None):
                    if master_agent['acct_status'] == 'Active':
                        result[master_agent['master_agent']] += 1
                else:
                    result[master_agent['master_agent']] = 1
            for master_agent, enroll_count in result.items():
                list_of_inserts.append({
                    'date':target_timestamp,
                    'name':master_agent,
                    'count':enroll_count,
                    'type':'Active Master Agent'
                })


            stmt = insert(Master_Agent_And_Agent_Active_Counts)
            conn.execute(stmt, list_of_inserts)
            conn.commit()


def populate_master_agent_total_count():
    with engine.connect() as conn:
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'MASTER_AGENT_NAME'])
                                           .where(Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['MASTER_AGENT_NAME'] != '')).all()
        main_table_contacts = {
            contact[0]:contact[1] for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        for date_delta in range(1,33):
            target_timestamp = time.mktime((datetime.today() - timedelta(days=date_delta)).date().timetuple())
            print(f"Starting day: {datetime.fromtimestamp(target_timestamp).strftime('%m/%d/%Y')}")
            new_contacts = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID'])
                                        .where(Merged_Database_Changes.c['Field_Name'] == 'New Contacts', 
                                               Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for new_contact in new_contacts:
                if main_table_contacts.get(new_contact[0], None):
                    del main_table_contacts[new_contact[0]]
            
            agent_changes = conn.execute(select(Merged_Database_Changes.c['Customer_Order_ID', 'Old_Value'])
                                      .where(Merged_Database_Changes.c['Field_Name'] == 'MASTER_AGENT_NAME',
                                             Merged_Database_Changes.c['Change_Date'] == target_timestamp)).all()
            for contact in agent_changes:
                if main_table_contacts.get(contact[0], None):
                    main_table_contacts[contact[0]] = contact[1]

            
            result = {}
            list_of_inserts = []
            for master_agent in main_table_contacts.values():
                if result.get(master_agent, None):
                    result[master_agent] += 1
                else:
                    result[master_agent] = 1
            for master_agent, enroll_count in result.items():
                list_of_inserts.append({
                    'date':target_timestamp,
                    'name':master_agent,
                    'count':enroll_count,
                    'type':'Total Master Agent'
                })


            stmt = insert(Master_Agent_And_Agent_Active_Counts)
            conn.execute(stmt, list_of_inserts)
            conn.commit()

def create_json_for_local_testing():
    with engine.connect() as conn, open('master_agent_agent_data.json', 'w') as json_file:
        data = conn.execute(select(Master_Agent_And_Agent_Active_Counts.c['id', 'date', 'name', 'count', 'type'])).all()
        result = {}
        for piece in data:
            result[piece[0]] = {
                'date':piece[1],
                'name':piece[2],
                'count':piece[3],
                'type':piece[4]
            }
        json.dump(result, json_file)

def making_address_report_for_ronna():
    with engine.connect() as conn:
        contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'First_Name', 'Last_Name', 'Address1', 'City', 'State', 'Zip', 'Activation_Date'])
                             .where(Merged_Customer_Databases.c['Source_Database'] == 'Telgoo',
                                    Merged_Customer_Databases.c['Activation_Date'] > datetime(year=2023, month=12, day=2).timestamp(),
                                    Merged_Customer_Databases.c['Activation_Date'] < datetime(year=2024, month=5, day=22).timestamp(),
                                    Merged_Customer_Databases.c['Activation_Date'] != '',
                                    Merged_Customer_Databases.c['Address1'] != '',
                                    Merged_Customer_Databases.c['City'] != '',
                                    Merged_Customer_Databases.c['State'] != '')
                                    .order_by(Merged_Customer_Databases.c['Address1'])).all()
    contacts = {c[0]:c for c in contacts}
    results = {}
    for contact in contacts.values():
        acs = ','.join([contact[3], contact[4], contact[5]])
        if results.get(acs, None):
            results[acs].append(contact[0])
        else:
            results[acs] = [contact[0]]
            
    with Workbook('report.xlsx') as wb:
        ws_one = wb.add_worksheet('Report One')
        ws_one.write_row(0,0,['Enrollment ID', 'First Name', 'Last Name', 'Address', 'City', 'State', 'Zip', 'Activation Date'])
        counter = 1
        for v in results.values():
            if len(v) > 1:
                for id in v:
                    row_data = [x for x in contacts[id]]
                    row_data[7] = datetime.fromtimestamp(row_data[7]).strftime('%m/%d/%Y')
                    ws_one.write_row(counter, 0, row_data)
                    counter += 1
        ws_two = wb.add_worksheet('Report Two')
        ws_two.write_row(0,0,['Address', 'City', 'State', 'Zip', '# of times address appears', 'Earliest Activation Date', 'Last Activation Date'])
        counter = 1
        for acs, ids in results.items():
            if len(ids) > 1:
                address, city, state = acs.split(',')
                zips = set()
                activation_dates = []
                for id in ids:
                    zips.add(contacts[id][6])
                    activation_dates.append(contacts[id][7])
                zips = ', '.join(zips)
                activation_dates.sort(reverse=True)
                earliest_date = datetime.fromtimestamp(activation_dates[-1]).strftime('%m/%d/%Y')
                latest_date = datetime.fromtimestamp(activation_dates[0]).strftime('%m/%d/%Y')
                ws_two.write_row(counter, 0, [address, city, state, zips, len(ids), earliest_date, latest_date])
                counter += 1



def audit():
    wb = load_workbook('/home/ubuntu/MaxsipReports/pythonFiles/Duplicate audit summary.xlsx')
    ws = wb.active
    contact_data = {}
    for row_num, row in enumerate(ws.rows):
        row_data = [x.value for x in row]
        if row_num == 0:
            fields = row_data
            continue
        contact_data[str(row_data[2])] = row_data
        
    enroll_ids_for_query = []
    with engine.connect() as conn:
        for enroll_id in contact_data.keys():
            enroll_ids_for_query.append(enroll_id)
            if len(enroll_ids_for_query) == 950:
                data_to_add = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'First_Name', 'Last_Name', 'DOB', 'SSN'])
                                           .where(Merged_Customer_Databases.c['Customer_Order_ID'].in_(enroll_ids_for_query))).all()
                
                for data in data_to_add:
                    data = [x for x in data]
                    if data[3]:
                        data[3] = datetime.fromtimestamp(data[3]).strftime('%m/%d/%Y')
                    contact_data[data[0]] += [data[1], data[2], data[3], data[4]]
                enroll_ids_for_query = []
        if enroll_ids_for_query:
            data_to_add = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'First_Name', 'Last_Name', 'DOB', 'SSN'])
                                        .where(Merged_Customer_Databases.c['Customer_Order_ID'].in_(enroll_ids_for_query))).all()
            for data in data_to_add:
                data = [x for x in data]
                if data[3]:
                    data[3] = datetime.fromtimestamp(data[3]).strftime('%m/%d/%Y')
                contact_data[data[0]] += [data[1], data[2], data[3], data[4]]
    with Workbook('Audit Update.xlsx') as wb:
        ws = wb.add_worksheet()
        ws.write_row(0,0,fields + ['First_Name', 'Last_Name', 'DOB', 'SSN'])
        counter = 1
        for val in contact_data.values():
            ws.write_row(counter, 0, val)
            counter += 1
        

def add_tracking_record_to_status_report():
    with engine.connect() as conn:
        yesterday_timestamp = time.mktime((datetime.today() - timedelta(days=1)).date().timetuple())
        conn.execute(insert(Status_Report_Counts)
                     .values(
                         date=yesterday_timestamp,
                         field='Last Updated',
                         count=None
                     ))
        conn.commit()


# INCOMPLETE
def add_new_contacts_to_groups():
    path_to_db = 'sqlite:////home/ubuntu/MaxsipReports/instance/site.db'
    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)

    qualifications_to_be_a_duplicate = {
        0:['NLAD_Subscriber_ID'],
        1:['First_Name', 'Last_Name', 'Address1', 'Zip'],
        2:['DOB', 'SSN', 'First_Name'],
        3:['DOB', 'SSN', 'Last_Name'],
        4:['SSN', 'First_Name', 'Address1'],
        5:['SSN', 'Last_Name', 'Address1'],
        6:['DOB', 'First_Name', 'Address1'],
        7:['DOB', 'Last_Name', 'Address1']
    }

    stmt = select(Merged_Customer_Databases.c['Customer_Order_ID', 'NLAD_Subscriber_ID', 'DOB', 
                                                'SSN', 'First_Name', 'Last_Name', 'Address1', 'Zip', 'Group_ID'])
    fields_dict = {k:v for v, k in enumerate(stmt.selected_columns.keys())}
    
    with engine.connect() as conn:
        contacts = conn.execute(stmt).all()
    
    total_contacts_length = len(contacts)


    def check_for_falsy(contact:list) -> bool:
        for data_piece in contact[:fields_dict['Group_ID']]:
            if bool(data_piece) == False:
                return False
        return True
    
    contacts_with_missing_data = {}

    customer_order_id_to_group = {}

    contacts_to_be_grouped_unfiltered_for_missing_data = []
    
    contacts_to_be_grouped_filtered = []

    list_of_contacts_to_be_updated_as_having_missing_data = []

    for contact in contacts:
        if contact[fields_dict['Group_ID']] == 1:
            contacts_with_missing_data[contact[fields_dict['Customer_Order_ID']]] = contact
        elif contact[fields_dict['Group_ID']]:
             customer_order_id_to_group[contact[fields_dict['Customer_Order_ID']]] = contact[fields_dict['Group_ID']]
        else:
            contacts_to_be_grouped_unfiltered_for_missing_data.append(contact)
    print(f'Contacts to be grouped before filter: {len(contacts_to_be_grouped_unfiltered_for_missing_data)}')

    for contact in contacts_to_be_grouped_unfiltered_for_missing_data:
        are_all_data_pieces_truthy = check_for_falsy(contact)
        if are_all_data_pieces_truthy:
            contacts_to_be_grouped_filtered.append(contact)
        else:
            contacts_with_missing_data[contact[fields_dict['Customer_Order_ID']]] = contact
            list_of_contacts_to_be_updated_as_having_missing_data.append({
                'b_Customer_Order_ID':contact[fields_dict['Customer_Order_ID']],
                'Group_ID':1
            })
    with engine.connect() as conn:
        conn.execute(update(Merged_Customer_Databases)
                     .where(Merged_Customer_Databases.c['Customer_Order_ID'] == bindparam('b_Customer_Order_ID'))
                     .values(Group_ID = bindparam('Group_ID'))
                     ,list_of_contacts_to_be_updated_as_having_missing_data)
        conn.commit()
    print(f'Contacts to be grouped after filter: {len(contacts_to_be_grouped_filtered)}')
    

    order_id_indexed_bank = {contact[fields_dict['Customer_Order_ID']]:contact for contact in contacts}

    total_bank = {}
    for params in qualifications_to_be_a_duplicate.values():
        total_bank['.'.join(params)] = {}
    for joined_params_names, param_values_to_order_id_list in total_bank.items():
        split_params = joined_params_names.split('.')
        print(f'starting: {split_params}')
        for contact in contacts:
            customer_order_id = contact[fields_dict['Customer_Order_ID']]
            if contacts_with_missing_data.get(customer_order_id, None):
                continue
            joined_param_values_of_contact = '.'.join([str(contact[fields_dict[param]]) for param in split_params])
            if param_values_to_order_id_list.get(joined_param_values_of_contact, None):
                param_values_to_order_id_list[joined_param_values_of_contact].append(customer_order_id)
            else:
                param_values_to_order_id_list[joined_param_values_of_contact] = [customer_order_id]
        print(f'Length of keys: {len(param_values_to_order_id_list.keys())}\nRemaining memory: {psutil.virtual_memory().available}')


    with engine.connect() as conn:
        new_max_group_num = conn.execute(select(func.max(Merged_Customer_Databases.c['Group_ID']))).first()[0] + 1

    def get_matches_by_params(total_group:list, params:list) -> list:
        matches = []
        for single_contact in total_group:
            joined_contact_params = '.'.join([str(single_contact[fields_dict[param]]) for param in params])
            list_of_common_order_ids = total_bank['.'.join(params)][joined_contact_params]
            for order_id_match in list_of_common_order_ids:
                if not customer_order_id_to_group.get(order_id_match, None):
                    matches.append(order_id_indexed_bank[order_id_match])
        return matches

    def go_through_data_sets(contact_data:list, index:int, match_origin:str='Parent'):
        total_group = [contact_data]
        if index == 8:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': new_max_group_num,
                'Match ID':match_origin,
                'Matched Params':''
            }
        else:
            customer_order_id_to_group[contact_data[fields_dict['Customer_Order_ID']]] = {
                'Group Number': new_max_group_num,
                'Match ID':match_origin,
                'Matched Params':index + 1
            }
        for number, param_set in qualifications_to_be_a_duplicate.items():
            if number == index:
                break
            matches = get_matches_by_params(total_group, param_set)
            if matches:
                total_group += matches
                for contact_match in matches:
                    go_through_data_sets(contact_match, number, contact_data[fields_dict['Customer_Order_ID']])



    for contact in contacts_to_be_grouped_filtered:
        customer_order_id = contact[fields_dict['Customer_Order_ID']]
        if customer_order_id_to_group.get(customer_order_id, None) or contacts_with_missing_data.get(customer_order_id, None):
            continue
        row_data = [x for x in contact]
        for number, param_set in qualifications_to_be_a_duplicate.items():
            joined_contact_params = '.'.join([str(contact[fields_dict[param]]) for param in param_set])
            list_of_common_order_ids = total_bank['.'.join(param_set)][joined_contact_params]
            for order_id in list_of_common_order_ids:
                if order_id != customer_order_id:
                    pass


        new_max_group_num += 1

def enrollment_types_active_count():
    engine = create_engine('sqlite:////home/ubuntu/MaxsipReports/instance/site.db')
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    Status_Report_Counts = Table("Status_Report_Counts", metadata_obj, autoload_with=engine)
    with engine.connect() as conn:
        today_timestamp = time.mktime(datetime.today().date().timetuple())
        main_table_contacts = conn.execute(select(Merged_Customer_Databases.c['Customer_Order_ID', 'ENROLLMENT_TYPE'])
                                           .where(Merged_Customer_Databases.c['Account_Status'] == 'Active',
                                                  Merged_Customer_Databases.c['Activation_Date'] != '',
                                                  Merged_Customer_Databases.c['ENROLLMENT_TYPE'] != '')).all()
        main_table_contacts = {
            contact[0]:contact[1] for contact in main_table_contacts
        }
        print(f'Selected contacts from main table, len: {len(main_table_contacts)}')

        result = {}
        list_of_inserts = []
        for e_type in main_table_contacts.values():
            if result.get(e_type, None):
                result[e_type] += 1
            else:
                result[e_type] = 1
        for e_type, enroll_count in result.items():
            list_of_inserts.append({
                'date':today_timestamp,
                'field':e_type,
                'count':enroll_count
            })


        stmt = insert(Status_Report_Counts)
        conn.execute(stmt, list_of_inserts)
        conn.commit()


def throwaway():
    timestart = datetime.now()
    with engine.connect() as conn:
        subquery = select(Merged_Customer_Databases.c.Plan.label('plan_name'), func.count().label('active_count')).select_from(Merged_Customer_Databases)\
            .where(Merged_Customer_Databases.c.Account_Status == 'Active')\
            .group_by(Merged_Customer_Databases.c.Plan).subquery()
        subquery_two = select(Merged_Customer_Databases.c.Plan.label('plan_name'), func.count().label('inactive_count')).select_from(Merged_Customer_Databases)\
            .where(Merged_Customer_Databases.c.Account_Status != 'Active')\
            .group_by(Merged_Customer_Databases.c.Plan).subquery()
        
        query = select(
            Plan_Assignment.c.id,
            Plan_Assignment.c.plan_name,
            subquery.c.active_count,
            subquery_two.c.inactive_count,
            Plan_Assignment.c.program_assignment,
            ).select_from(
                Plan_Assignment
                .outerjoin(subquery, Plan_Assignment.c.plan_name == subquery.c.plan_name)
                .outerjoin(subquery_two, Plan_Assignment.c.plan_name == subquery_two.c.plan_name)
            ).where(
                Plan_Assignment.c.source_database == 'Unavo'
            ).order_by(
                Plan_Assignment.c.id
            ).limit(
                5
            )
        plans = conn.execute(query).all()
        result = []
        for plan in plans:
            holding = []
            for x in plan:
                if x == None:
                    x = 0
                holding.append(x)
            result.append(holding)
    print(f'Total time for request: {datetime.now() - timestart}')

def throwaway_two():
    timestart = datetime.now()
    with engine.connect() as conn:
        # Main query with conditional aggregation
        query = select(
            Plan_Assignment.c.id,
            Plan_Assignment.c.plan_name,
            func.count(case((Merged_Customer_Databases.c.Account_Status == 'Active', 1), else_=None)).label('active_count'),
            func.count(case((Merged_Customer_Databases.c.Account_Status != 'Active', 1), else_=None)).label('inactive_count'),
            Plan_Assignment.c.program_assignment
        ).select_from(
            Plan_Assignment
            .outerjoin(Merged_Customer_Databases, Plan_Assignment.c.plan_name == Merged_Customer_Databases.c.Plan)
        ).where(
            Plan_Assignment.c.source_database == 'Unavo'
        ).group_by(
            Plan_Assignment.c.id,
            Plan_Assignment.c.plan_name,
            Plan_Assignment.c.program_assignment
        ).order_by(
            Plan_Assignment.c.id
        ).limit(5)

        # Execute the query
        result = conn.execute(query).all()
        for x in result:
            print(x)

    print(f'Total time for request: {datetime.now() - timestart}')
        
def throwthree():
    with engine.connect() as conn:
        plans = conn.execute(select(Plan_Assignment.c['id', 'plan_name', 'source_database', 'program_assignment'])).all()
    
    with open('plan_assignments.csv', 'w') as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerow(['ID', 'Plan Name', 'Source Database', 'Program Assignment'])
        csvwriter.writerows(plans)


if __name__ == '__main__':
    path_to_db = 'sqlite:////home/ubuntu/MaxsipReports/instance/site.db'

    engine = create_engine(path_to_db)
    metadata_obj = MetaData()
    metadata_obj.reflect(bind=engine)
    unavoChanges = Table("Unavo_Changes", metadata_obj, autoload_with=engine)
    telgooChanges = Table("Telgoo_Changes", metadata_obj, autoload_with=engine)
    customerInfoTelgoo = Table("customerInfoTelgoo", metadata_obj, autoload_with=engine)
    cdd = Table("Customer_Device_Database", metadata_obj, autoload_with=engine)
    main_customer_view = Table("Main_Customer_Database", metadata_obj, autoload_with=engine)
    NLAD_Subscriber_ID_Count_View = Table("NLAD_Subscriber_ID_Count_View", metadata_obj, autoload_with=engine)
    customerInfo = Table("customerInfo", metadata_obj, autoload_with=engine)
    deviceQty = Table('Device_Qty', metadata_obj, autoload_with=engine)
    AgentSalesReports = Table('AgentSalesReports', metadata_obj, autoload_with=engine)
    Merged_Customer_Databases = Table("Merged_Customer_Databases", metadata_obj, autoload_with=engine)
    Merged_Database_Changes = Table("Merged_Database_Changes", metadata_obj, autoload_with=engine)
    Status_Report_Counts = Table("Status_Report_Counts", metadata_obj, autoload_with=engine)
    Master_Agent_And_Agent_Active_Counts = Table("Master_Agent_And_Agent_Active_Counts", metadata_obj, autoload_with=engine)
    Plan_Assignment = Table("Plan_Assignment", metadata_obj, autoload_with=engine)

    yesterday = (datetime.today() - timedelta(days=1)).date()
    yesterday_timestamp = time.mktime((datetime.today() - timedelta(days=1)).date().timetuple())
